"""The safety adviser's annual report: a statistic over the kept shipments.

ADR 1.8.3.3 obliges the adviser to draw up a yearly report to management on
the undertaking's activities in the carriage of dangerous goods, kept five
years and shown to the authorities on request. The provision prescribes the
report and not its contents, so this module does two different things and
keeps them apart:

- **counts what the history can prove** — how many shipments, with what,
  where to, in which class and quantity, and what the 1.1.3.6 points count
  said about each — over the shipments kept in one calendar year that the
  person drawing up the report may see (the same rule the shipments page
  uses, ``departments.visible_to``);
- **offers the adviser's own duties as headings** — the list 1.8.3.3 gives
  of the practices the adviser must check — with nothing filled in. A
  generated opinion on training or emergency procedures would be worse than
  a blank, so the blank is deliberate.

A shipment that was never kept is not counted, and the report says so.
Quantities are the declared totals per substance as the export holds them,
read with the same parser 1.1.3.6 uses, so kilograms and litres stay apart
and a quantity without a unit is counted as unknown rather than guessed.
"""
from __future__ import annotations

import io
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from functools import lru_cache
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import extract
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.languages import normalise, pick
from app.models.shipment import Shipment
from app.models.user import Department, User
from app.services import departments
from app.services.dg.autofill import _num, adr_quantity
from app.services.documents.registry import get_registry
from app.version import get_version

#: The 1.1.3.6 outcomes the compliance check can give, in the order the
#: report lists them. Anything else the export holds is listed after these.
POINTS_STATUSES = ("exempt_possible", "above_threshold", "not_exempt",
                   "incomplete", "not_available_for_mode")


@lru_cache
def texts() -> dict[str, Any]:
    path = get_settings().config_dir / "dgsa_report.json"
    return json.loads(path.read_text(encoding="utf-8"))


def _loads(raw: str | None) -> dict[str, Any]:
    try:
        value = json.loads(raw or "{}")
    except ValueError:
        return {}
    return value if isinstance(value, dict) else {}


def _un(product: dict[str, Any]) -> str:
    digits = "".join(c for c in str(product.get("un_number") or "") if c.isdigit())
    return digits.zfill(4) if digits else ""


def _class(product: dict[str, Any]) -> str:
    return str(product.get("class") or "").strip()


def years_kept(db: Session, viewer: User) -> list[int]:
    """The calendar years with at least one kept shipment the viewer may see,
    newest first — the choices the page offers."""
    query = departments.visible_to(db.query(extract("year", Shipment.created_at)), viewer)
    found = {int(row[0]) for row in query.distinct().all() if row[0] is not None}
    return sorted(found, reverse=True)


def _modality_labels() -> dict[str, dict[str, str]]:
    return {m["key"]: m.get("label", {}) for m in get_registry().get("modalities", [])}


def _document_labels() -> dict[str, dict[str, str]]:
    return {d["key"]: d.get("label", {}) for d in get_registry().get("documents", [])}


def build_report(db: Session, viewer: User, year: int, department: str = "",
                 language: str = "nl") -> dict[str, Any]:
    """The report for one calendar year, as the page and the workbook read it."""
    lang = normalise(language)
    labels = texts()["labels"]
    L = lambda key: pick(labels.get(key), lang, key)  # noqa: E731

    query = departments.visible_to(db.query(Shipment), viewer, department)
    rows = (query.filter(extract("year", Shipment.created_at) == year)
            .order_by(Shipment.created_at.asc(), Shipment.id.asc()).all())

    months = [{"month": m, "shipments": 0, "with_dangerous_goods": 0} for m in range(1, 13)]
    by_modality: dict[str, dict[str, int]] = defaultdict(lambda: {"shipments": 0, "with_dangerous_goods": 0})
    by_regulation: Counter[str] = Counter()
    by_department: dict[str, dict[str, int]] = defaultdict(lambda: {"shipments": 0, "with_dangerous_goods": 0})
    by_class: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"shipments": set(), "products": 0, "quantity_kg": 0.0, "quantity_l": 0.0, "quantity_unknown": 0})
    by_un: dict[tuple[str, str, str], dict[str, Any]] = defaultdict(
        lambda: {"name": "", "shipments": set(), "products": 0,
                 "quantity_kg": 0.0, "quantity_l": 0.0, "quantity_unknown": 0})
    points: Counter[str] = Counter()
    documents: Counter[str] = Counter()
    carriage_modes: Counter[str] = Counter()
    hcdg: dict[str, dict[str, Any]] = {}
    class7_packages = 0
    with_dg = 0
    products_total = 0
    kg_total = l_total = 0.0
    unknown_total = 0

    for record in rows:
        export = _loads(record.export_json)
        entries = export.get("dangerous_goods") or []
        has_dg = bool(entries)
        with_dg += int(has_dg)
        month = months[record.created_at.month - 1]
        month["shipments"] += 1
        month["with_dangerous_goods"] += int(has_dg)
        modality = by_modality[record.modality or ""]
        modality["shipments"] += 1
        modality["with_dangerous_goods"] += int(has_dg)
        dept_name = record.department.name if record.department else ""
        dept = by_department[dept_name]
        dept["shipments"] += 1
        dept["with_dangerous_goods"] += int(has_dg)
        for regime in (record.regulations or "").split(","):
            if regime:
                by_regulation[regime] += 1
        for key in export.get("documents") or []:
            documents[str(key)] += 1
        if has_dg:
            compliance = export.get("compliance") or {}
            status = str((compliance.get("adr_points") or {}).get("status") or "")
            points[status or "not_assessed"] += 1
            # 1.10.3: which high consequence dangerous goods were consigned,
            # as the check found them when the shipment was kept. One line
            # per UN number, the first reason kept.
            for item in (compliance.get("adr_security") or {}).get("items") or []:
                if not isinstance(item, dict) or item.get("not_answered") or not item.get("un_number"):
                    continue
                un = str(item["un_number"]).strip()
                entry = hcdg.setdefault(un, {"un_number": un, "reason": str(item.get("reason") or ""),
                                             "carriage_mode": str(item.get("carriage_mode") or ""),
                                             "shipments": set()})
                entry["shipments"].add(record.id)
        for entry in entries:
            for product in entry.get("products") or []:
                if not isinstance(product, dict):
                    continue
                products_total += 1
                cls = _class(product) or "?"
                # The wizard says "packages", "tank", "portable_tank" or "bulk";
                # the form asks package / tank / bulk.
                mode = str(product.get("carriage_mode") or "").strip() or "packages"
                carriage_modes[{"packages": "package", "portable_tank": "tank"}.get(mode, mode)] += 1
                if cls.split(".")[0] == "7":
                    packages = _num(product.get("quantity_packages"))
                    class7_packages += int(packages) if packages else 0
                value, unit = adr_quantity(product)
                bucket = by_class[cls]
                key = (_un(product), cls, str(product.get("packing_group") or "").strip())
                per_un = by_un[key]
                if not per_un["name"]:
                    per_un["name"] = str(product.get("proper_shipping_name") or "").strip()
                for target in (bucket, per_un):
                    target["shipments"].add(record.id)
                    target["products"] += 1
                    if value is None or not unit:
                        target["quantity_unknown"] += 1
                    elif unit == "kg":
                        target["quantity_kg"] += value
                    else:
                        target["quantity_l"] += value
                if value is None or not unit:
                    unknown_total += 1
                elif unit == "kg":
                    kg_total += value
                else:
                    l_total += value

    def _finish(bucket: dict[str, Any]) -> dict[str, Any]:
        return {**bucket, "shipments": len(bucket["shipments"]),
                "quantity_kg": round(bucket["quantity_kg"], 3),
                "quantity_l": round(bucket["quantity_l"], 3)}

    def _class_sort(cls: str) -> tuple[int, str]:
        head = cls.split(".")[0]
        return (int(head) if head.isdigit() else 99, cls)

    modality_labels = _modality_labels()
    document_labels = _document_labels()
    if viewer.role == "admin":
        if department == "none":
            scope = L("scope_none")
        elif department.isdigit():
            found = db.get(Department, int(department))
            scope = found.name if found else L("scope_all")
        else:
            scope = L("scope_all")
    else:
        scope = (viewer.department.name if getattr(viewer, "department", None) else L("scope_none"))

    status_rows = [{"status": s, "label": L(f"status_{s}"), "shipments": points[s]}
                   for s in POINTS_STATUSES if points[s]]
    status_rows += [{"status": s, "label": L("status_not_assessed") if s == "not_assessed" else s,
                     "shipments": n} for s, n in points.items() if s not in POINTS_STATUSES]

    return {
        "year": year,
        "language": lang,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "generated_by": viewer.username,
        "generator": {"application": "CargoPilot", "version": get_version()},
        "scope": scope,
        "basis": pick(texts()["basis"], lang),
        "source": texts()["source"],
        "counted_note": pick(texts()["counted_note"], lang),
        "totals": {
            "shipments": len(rows),
            "with_dangerous_goods": with_dg,
            "without_dangerous_goods": len(rows) - with_dg,
            "products": products_total,
            "quantity_kg": round(kg_total, 3),
            "quantity_l": round(l_total, 3),
            "quantity_unknown": unknown_total,
        },
        "by_month": months,
        "by_modality": [{"modality": key, "label": pick(modality_labels.get(key), lang, key or L("unknown")), **v}
                        for key, v in sorted(by_modality.items())],
        "by_regulation": [{"regulation": key, "shipments": n} for key, n in sorted(by_regulation.items())],
        "by_department": [{"department": key or L("scope_none"), **v}
                          for key, v in sorted(by_department.items(), key=lambda kv: (kv[0] == "", kv[0]))],
        "by_class": [{"class": cls, **_finish(v)} for cls, v in sorted(by_class.items(), key=lambda kv: _class_sort(kv[0]))],
        "by_un_number": [
            {"un_number": un, "class": cls, "packing_group": pg, **_finish(v)}
            for (un, cls, pg), v in sorted(by_un.items(), key=lambda kv: (-len(kv[1]["shipments"]), kv[0]))],
        "adr_points": status_rows,
        "carriage_modes": [mode for mode, _n in sorted(carriage_modes.items())],
        "high_consequence": [{**v, "shipments": len(v["shipments"])}
                             for _un, v in sorted(hcdg.items())],
        "class7_packages": class7_packages,
        "documents": [{"document": key, "label": pick(document_labels.get(key), lang, key), "shipments": n}
                      for key, n in sorted(documents.items(), key=lambda kv: (-kv[1], kv[0]))],
        "duties_heading": pick(texts()["duties_heading"], lang),
        "duties": [{"key": d["key"], "text": pick(d, lang)} for d in texts()["duties"]],
    }


# --- the workbook -----------------------------------------------------------


def build_workbook(report: dict[str, Any]) -> bytes:
    """The report as an .xlsx: one sheet per table, the duties last with an
    empty column for the adviser's finding."""
    lang = report["language"]
    labels = texts()["labels"]
    L = lambda key: pick(labels.get(key), lang, key)  # noqa: E731
    bold = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = L("summary")[:31]
    ws.append([L("title")])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([L("year"), report["year"]])
    ws.append([L("scope"), report["scope"]])
    ws.append([L("generated_at"), report["generated_at"]])
    ws.append([L("generated_by"), report["generated_by"]])
    ws.append([])
    ws.append([report["basis"]])
    ws.append([report["counted_note"]])
    ws.append([])
    totals = report["totals"]
    for key in ("shipments", "with_dangerous_goods", "without_dangerous_goods", "products",
                "quantity_kg", "quantity_l", "quantity_unknown"):
        ws.append([L(key), totals[key]])
    ws.append([])
    ws.append([L("adr_points")])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([L("adr_points_note")])
    for row in report["adr_points"]:
        ws.append([row["label"], row["shipments"]])
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 16

    def sheet(title: str, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
        page = wb.create_sheet(title[:31])
        page.append(headers)
        for cell in page[1]:
            cell.font = bold
        for row in rows:
            page.append(row)
        for index, width in enumerate(widths, start=1):
            page.column_dimensions[chr(64 + index)].width = width

    sheet(L("by_month"), [L("month"), L("shipments"), L("with_dangerous_goods")],
          [[m["month"], m["shipments"], m["with_dangerous_goods"]] for m in report["by_month"]], [10, 14, 24])
    sheet(L("by_modality"), [L("modality"), L("shipments"), L("with_dangerous_goods")],
          [[m["label"], m["shipments"], m["with_dangerous_goods"]] for m in report["by_modality"]], [28, 14, 24])
    sheet(L("by_regulation"), [L("regulation"), L("shipments")],
          [[r["regulation"], r["shipments"]] for r in report["by_regulation"]], [16, 14])
    sheet(L("by_department"), [L("department"), L("shipments"), L("with_dangerous_goods")],
          [[d["department"], d["shipments"], d["with_dangerous_goods"]] for d in report["by_department"]], [28, 14, 24])
    sheet(L("by_class"), [L("class"), L("shipments"), L("products"), L("quantity_kg"), L("quantity_l"), L("quantity_unknown")],
          [[c["class"], c["shipments"], c["products"], c["quantity_kg"], c["quantity_l"], c["quantity_unknown"]]
           for c in report["by_class"]], [10, 14, 16, 16, 16, 26])
    sheet(L("by_un_number"),
          [L("un_number"), L("name"), L("class"), L("packing_group"), L("shipments"), L("products"),
           L("quantity_kg"), L("quantity_l"), L("quantity_unknown")],
          [[u["un_number"], u["name"], u["class"], u["packing_group"], u["shipments"], u["products"],
            u["quantity_kg"], u["quantity_l"], u["quantity_unknown"]] for u in report["by_un_number"]],
          [12, 44, 8, 14, 14, 16, 16, 16, 26])
    sheet(L("documents"), [L("document"), L("shipments")],
          [[d["label"], d["shipments"]] for d in report["documents"]], [40, 14])
    sheet(L("duties"), [L("duty"), L("finding")],
          [[d["text"], ""] for d in report["duties"]], [100, 60])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
