"""Reading and writing import templates (xlsx, csv, txt)."""

from __future__ import annotations

import csv
import io
import zipfile
from typing import Any

import openpyxl
from openpyxl import Workbook

from app.core.messages import text as message_text

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000
MAX_IMPORT_COLUMNS = 100
MAX_IMPORT_CELL_CHARS = 10_000


class ImportLimitError(ValueError):
    """The import exceeds an explicit safety limit.

    Carries the message code as well as the English text, so the route that
    turns it into an HTTP error can hand the interface something translatable
    instead of a sentence in one language.
    """

    def __init__(self, code: str, **params: Any) -> None:
        super().__init__(message_text(code, **params))
        self.code = code
        self.params = params


async def read_limited_upload(file: Any, max_bytes: int = MAX_IMPORT_BYTES) -> bytes:
    """Never read more than the configured limit plus one detection byte."""
    content = await file.read(max_bytes + 1)
    if len(content) > max_bytes:
        raise ImportLimitError("import.file_too_large", limit_mb=max_bytes // (1024 * 1024))
    return content


def build_xlsx(headers: list[str], rows: list[list[str]], sheet_name: str = "Export") -> bytes:
    """A workbook of one sheet: bold headers, then the rows as given."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_xlsx_template(headers: list[str], example_row: list[str], sheet_name: str = "Template") -> bytes:
    return build_xlsx(headers, [example_row], sheet_name=sheet_name)


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _split_line(line: str) -> list[str]:
    line = line.strip()
    if not line:
        return []
    if "\t" in line:
        return [c.strip() for c in line.split("\t")]
    if "|" in line:
        return [c.strip() for c in line.split("|")]
    if ";" in line and line.count(";") >= 1:
        return [c.strip() for c in line.split(";")]
    if "," in line and line.count(",") >= 1:
        return [c.strip() for c in line.split(",")]
    return [line]


def _check_row(
    cells: list[str],
    row_number: int,
    max_columns: int | None,
    max_cell_chars: int | None,
) -> None:
    if max_columns is not None and len(cells) > max_columns:
        raise ImportLimitError(
            "import.too_many_columns",
            row=row_number,
            columns=len(cells),
            limit=max_columns,
        )
    if max_cell_chars is not None:
        for column_number, cell in enumerate(cells, start=1):
            if len(cell) > max_cell_chars:
                raise ImportLimitError(
                    "import.cell_too_long",
                    row=row_number,
                    column=column_number,
                    limit=max_cell_chars,
                )


def validate_tabular_rows(
    rows: list[list[str]],
    *,
    max_rows: int = MAX_IMPORT_ROWS,
    max_columns: int = MAX_IMPORT_COLUMNS,
    max_cell_chars: int = MAX_IMPORT_CELL_CHARS,
) -> None:
    if len(rows) > max_rows:
        raise ImportLimitError("import.too_many_rows", rows=len(rows), limit=max_rows)
    for row_number, row in enumerate(rows, start=1):
        _check_row(row, row_number, max_columns, max_cell_chars)


def _append_checked_row(
    rows: list[list[str]],
    cells: list[str],
    *,
    max_rows: int | None,
    max_columns: int | None,
    max_cell_chars: int | None,
) -> None:
    if not any(cells):
        return
    next_row = len(rows) + 1
    if max_rows is not None and next_row > max_rows:
        raise ImportLimitError("import.too_many_rows", rows=max_rows + 1, limit=max_rows)
    _check_row(cells, next_row, max_columns, max_cell_chars)
    rows.append(cells)


def read_tabular_file(
    content: bytes,
    filename: str,
    *,
    max_rows: int | None = None,
    max_columns: int | None = None,
    max_cell_chars: int | None = None,
) -> list[list[str]]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(content, max_rows, max_columns, max_cell_chars)
    if name.endswith(".csv"):
        return _read_csv(content, max_rows, max_columns, max_cell_chars)
    return _read_text(content, max_rows, max_columns, max_cell_chars)


def _validate_xlsx_archive(
    content: bytes,
    max_uncompressed_bytes: int = MAX_XLSX_UNCOMPRESSED_BYTES,
) -> None:
    """Reject compressed spreadsheets that expand beyond the safe limit."""
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        uncompressed = sum(info.file_size for info in archive.infolist())
    if uncompressed > max_uncompressed_bytes:
        raise ImportLimitError(
            "import.unpacked_too_large",
            limit_mb=max_uncompressed_bytes // (1024 * 1024),
        )


def _read_xlsx(
    content: bytes,
    max_rows: int | None,
    max_columns: int | None,
    max_cell_chars: int | None,
) -> list[list[str]]:
    _validate_xlsx_archive(content)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_str(c) for c in row]
            _append_checked_row(
                rows,
                cells,
                max_rows=max_rows,
                max_columns=max_columns,
                max_cell_chars=max_cell_chars,
            )
        return rows
    finally:
        wb.close()


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv(
    content: bytes,
    max_rows: int | None,
    max_columns: int | None,
    max_cell_chars: int | None,
) -> list[list[str]]:
    text = _decode_text(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.reader(io.StringIO(text), dialect)
    rows: list[list[str]] = []
    for row in reader:
        cells = [c.strip() for c in row]
        _append_checked_row(
            rows,
            cells,
            max_rows=max_rows,
            max_columns=max_columns,
            max_cell_chars=max_cell_chars,
        )
    return rows


def _read_text(
    content: bytes,
    max_rows: int | None,
    max_columns: int | None,
    max_cell_chars: int | None,
) -> list[list[str]]:
    text = _decode_text(content)
    rows: list[list[str]] = []
    for line in text.splitlines():
        cells = _split_line(line)
        if cells:
            _append_checked_row(
                rows,
                cells,
                max_rows=max_rows,
                max_columns=max_columns,
                max_cell_chars=max_cell_chars,
            )
    return rows


def rows_to_pipe_text(rows: list[list[str]], skip_header: bool = False) -> str:
    start = 1 if skip_header and len(rows) > 1 else 0
    lines: list[str] = []
    for row in rows[start:]:
        while row and not row[-1]:
            row = row[:-1]
        if not any(cell.strip() for cell in row):
            continue
        lines.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(lines)


def normalize_header(value: str) -> str:
    return value.lower().strip().replace(" ", "_")
