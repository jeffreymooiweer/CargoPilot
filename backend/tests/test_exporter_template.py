import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.database import Base, SessionLocal, engine
from app.core.startup import ensure_directories, seed_catalogs
from app.services.exporter.appendix_exporter import export_appendix
from app.services.pipeline import parse_and_calculate

STEEL_INPUT = """Stalen hoekprofiel 80x80x8x6000 | 8 | stuks
UNP 220 mml=5700mm gegalvaniseerd | 24 | stuks"""


@pytest.fixture
def db(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{data_dir / 'test.db'}")
    get_settings.cache_clear()
    ensure_directories()
    template_src = Path(__file__).resolve().parents[2] / ".." / "templates" / "Appendix_A1D_template.xlsx"
    if template_src.exists():
        import shutil

        settings = get_settings()
        shutil.copy2(template_src, settings.templates_dir / "Appendix_A1D_template.xlsx")
    Base.metadata.create_all(bind=engine)
    session = SessionLocal()
    seed_catalogs(session)
    yield session
    session.close()
    get_settings.cache_clear()


def test_export_preserves_template_structure(db):
    settings = get_settings()
    template = settings.templates_dir / "Appendix_A1D_template.xlsx"
    if not template.exists():
        pytest.skip("Template not available")
    result = parse_and_calculate(STEEL_INPUT, db, output_language="en", mode="continue")
    out = export_appendix(result["lines"], {"route": "Test"}, "en", "testjob")
    assert out.exists()
    wb = load_workbook(out)
    assert " A1" in wb.sheetnames
    ws = wb[" A1"]
    assert ws["B12"].value == "Construction materials"
    assert ws["E12"].value == 8
    assert ws["N12"].value is not None
