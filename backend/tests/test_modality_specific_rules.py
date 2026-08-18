"""What applies to the road does not automatically apply to rail or water.

CargoPilot treated ADR, RID and ADN as one thing. That is understandable — the
three regimes resemble each other closely and share their structure — but it
produces two kinds of fault, and the second is the worse one.

**An invented entry on an official document.** The tunnel restriction code comes
from column 15 of ADR Table A and belongs on the road document under 5.4.1.1.1
(k). RID Table A does not have that column and the ADN transport document does
not carry it either. Yet the app put "(D/E)" on a CIM waybill and on an ADN
document. That is not a missing check but incorrect information the app added
itself.

**An outcome that looks stricter or looser than it is.** The 1.1.3.6 points and
the mixed loading of 7.5.2 are computed with the ADR tables. RID and ADN have
their own versions of those chapters, and those are not in CargoPilot. Presenting
the outcome silently as "the RID result" gives the user a certainty that does not
exist. They may have it as an indication — but then with that label on it.
"""

import pytest

from app.services.dg.autofill import description_line
from app.services.dg.compliance import check_compliance

GASOLINE = {
    "un_number": "1203",
    "proper_shipping_name": "GASOLINE",
    "class": "3",
    "packing_group": "II",
    "tunnel_code": "D/E",
    "transport_category": "2",
    "adr_total_quantity": "400",
    "quantity_packages": "10",
    "type_of_package": "jerrycan",
}


def entries():
    return [{"line_id": "L1", "products": [dict(GASOLINE)]}]


# --- The tunnel code belongs on the road document only --------------------


def test_the_road_document_carries_the_tunnel_code():
    assert "(D/E)" in description_line(dict(GASOLINE), "ADR")


@pytest.mark.parametrize("profile", ["RID", "ADN", "IMDG", "IATA_DGR"])
def test_no_other_document_carries_it(profile):
    """RID Table A has no column 15 with a tunnel code, and sea and air certainly
    do not. So nothing resembling it should appear in brackets."""
    assert "D/E" not in description_line(dict(GASOLINE), profile)


def test_the_cim_export_does_not_show_a_tunnel_code():
    """The CIM carries the RID profile; the goods column is what gets printed."""
    import openpyxl

    from app.services.documents.exporter import export_document
    from tests.test_documents import BASE_VALUES, LINES

    path = export_document("cim", dict(BASE_VALUES), LINES, entries(), language="nl")
    text = "\n".join(
        str(cell)
        for row in openpyxl.load_workbook(path).active.iter_rows(values_only=True)
        for cell in row
        if cell
    )
    assert "GASOLINE" in text, "de stof hoort er wel op te staan"
    assert "(D/E)" not in text


def test_the_cmr_export_does_show_one():
    import openpyxl

    from app.services.documents.exporter import export_document
    from tests.test_documents import BASE_VALUES, LINES

    path = export_document("cmr", dict(BASE_VALUES), LINES, entries(), language="nl")
    text = "\n".join(
        str(cell)
        for row in openpyxl.load_workbook(path).active.iter_rows(values_only=True)
        for cell in row
        if cell
    )
    assert "(D/E)" in text


# --- And the basis of a calculation is named ------------------------------


def test_a_road_shipment_gets_no_caveat():
    """For ADR the ADR table *is* the right table; there is nothing to add."""
    out = check_compliance(entries(), ["ADR"], "nl")
    assert out["adr_points"]["basis_note"] is None
    assert "adr_mixed_loading_basis_note" not in out


@pytest.mark.parametrize("profile", ["RID", "ADN"])
def test_rail_and_inland_waterway_say_which_tables_were_used(profile):
    """The basis is named — and since v1.33.0 with the text alongside.

    This test first required the note to contain "ADR", for both modes. That was
    right as long as both were computed with the ADR tables. After reading the
    official texts that is no longer true for ADN: ADN 1.1.3.6.1 has no points
    count and is assessed with its own table, so its note should precisely *not*
    refer to ADR. RID does compute the same way and now names its own article
    numbers.
    """
    out = check_compliance(entries(), [profile], "nl")
    note = out["adr_points"]["basis_note"]
    assert note and profile in note
    if profile == "RID":
        assert "1.1.3.6.3" in note and "ADR" in note
    else:
        assert "geen puntentelling" in note
        assert out["adn_exemption"]["basis"] == "ADN 1.1.3.6.1"
    # 7.5.2 is answered under the regime's own name — the rail table since
    # v1.41.0, the ADN's own 7.1.4 prohibitions since v1.119.0 — so a
    # single-regime selection needs no caveat about a table that was not used.
    assert "adr_mixed_loading_basis_note" not in out


def test_a_combined_selection_says_which_leg_the_7522_table_answers():
    """Road plus rail shows the road table, which additionally carries
    compatibility group A; road plus inland waterway shows the road outcome
    beside the ADN's own findings. Both differences are named."""
    out = check_compliance(entries(), ["ADR", "RID"], "nl")
    assert "groep A" in out["adr_mixed_loading_basis_note"]
    out = check_compliance(entries(), ["ADR", "ADN"], "nl")
    assert "7.1.4" in out["adr_mixed_loading_basis_note"]


def test_the_caveat_names_both_when_both_are_selected():
    out = check_compliance(entries(), ["ADR", "RID", "ADN"], "nl")
    note = out["adr_points"]["basis_note"]
    assert "RID" in note and "ADN" in note


def test_the_calculation_itself_is_unchanged():
    """The outcome stays the same — only the label is new. Otherwise this would
    quietly change the points count for road transport as well."""
    road = check_compliance(entries(), ["ADR"], "nl")["adr_points"]
    rail = check_compliance(entries(), ["RID"], "nl")["adr_points"]
    assert road["total_points"] == rail["total_points"] == 1200.0
    assert road["status"] == rail["status"]


def test_the_basis_is_named_in_every_language():
    for language in ("nl", "en", "de"):
        note = check_compliance(entries(), ["RID"], language)["adr_points"]["basis_note"]
        assert note and "ADR" in note and "RID" in note


def test_sea_and_air_never_get_the_land_checks_at_all():
    out = check_compliance(entries(), ["IMDG"], "nl")
    assert "adr_points" not in out
    assert "adr_mixed_loading" not in out
