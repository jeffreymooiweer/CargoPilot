import json
import os
import re
import shutil
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from app.core.config import get_settings


def sanitize_cell_value(value: Any) -> Any:
    if isinstance(value, str) and value and value[0] in "=+-@":
        return "'" + value
    return value


def load_mapping() -> dict[str, Any]:
    settings = get_settings()
    path = settings.config_dir / "appendix_mapping.json"
    return json.loads(path.read_text(encoding="utf-8"))


def _set_cell(ws, coord: str, value: Any) -> None:
    if value is None or value == "":
        return
    ws[coord] = sanitize_cell_value(value)


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int = 40) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy(src._style)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _shift_formula(formula: str, src_row: int, dst_row: int) -> str:
    if not formula or not formula.startswith("="):
        return formula
    delta = dst_row - src_row

    def repl(match: re.Match[str]) -> str:
        col = match.group(1)
        row = int(match.group(2)) + delta
        return f"{col}{row}"

    return re.sub(r"([A-Z]+)(\d+)", repl, formula)


def _copy_formulas(ws, source_row: int, target_row: int, formula_cols: list[str]) -> None:
    for col_letter in formula_cols:
        src_cell = ws[f"{col_letter}{source_row}"]
        dst_cell = ws[f"{col_letter}{target_row}"]
        if src_cell.value and str(src_cell.value).startswith("="):
            dst_cell.value = _shift_formula(str(src_cell.value), source_row, target_row)


def _yn(value: Any, default: str = "N") -> str:
    if value is None or value == "":
        return default
    text = str(value).strip().upper()
    if text in {"Y", "YES", "JA", "J", "TRUE", "1"}:
        return "Y"
    if text in {"N", "NO", "NEE", "FALSE", "0"}:
        return "N"
    return text


def _fill_sheet_d(
    wb,
    mapping: dict[str, Any],
    metadata: dict[str, Any],
    dangerous_goods: list[dict[str, Any]],
) -> None:
    sheet_d = mapping.get("sheet_d")
    if not sheet_d or not dangerous_goods:
        return
    ws = wb[sheet_d["sheet"]]
    meta = sheet_d["metadata"]
    if metadata.get("date"):
        _set_cell(ws, meta["date"], metadata["date"])
    else:
        _set_cell(ws, meta["date"], datetime.now().date())
    if metadata.get("route"):
        _set_cell(ws, meta["route"], metadata["route"])
    if metadata.get("ba_code"):
        _set_cell(ws, meta["ba_code"], metadata["ba_code"])

    field_rows = sheet_d["field_rows"]
    product_columns = sheet_d["product_columns"]

    for entry in dangerous_goods:
        vehicle = entry.get("vehicle") or entry.get("registration") or ""
        _set_cell(ws, meta["vehicle"], vehicle)
        products = entry.get("products") or []
        for product_index, product in enumerate(products[: len(product_columns)]):
            col = product_columns[product_index]
            for field_name, row_num in field_rows.items():
                value = product.get(field_name)
                if value is not None and value != "":
                    _set_cell(ws, f"{col}{row_num}", value)
        break


def export_appendix(
    lines: list[dict[str, Any]],
    metadata: dict[str, Any],
    output_language: str,
    job_ref: str,
    template_name: str | None = None,
    dangerous_goods: list[dict[str, Any]] | None = None,
) -> Path:
    settings = get_settings()
    mapping = load_mapping()
    template_file = template_name or mapping["template"]
    template_path = settings.templates_dir / template_file
    if not template_path.exists():
        bundled = Path(__file__).resolve().parents[3] / ".." / "templates" / template_file
        bundled = bundled.resolve()
        if bundled.exists():
            settings.templates_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(bundled, template_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb[mapping["sheet"]]
    cols = mapping["data"]["columns"]
    defaults = mapping["defaults"]
    start_row = mapping["data"]["start_row"]
    max_rows = mapping["data"]["max_rows"]
    style_row = mapping["data"]["style_source_row"]
    formula_cols = ["H", "I", "K", "M", "O", "P", "Q", "R", "S"]

    meta = mapping["metadata"]
    if metadata.get("date"):
        _set_cell(ws, meta["date"], metadata["date"])
    else:
        _set_cell(ws, meta["date"], datetime.now().date())
    if metadata.get("route"):
        _set_cell(ws, meta["route"], metadata["route"])
    if metadata.get("ba_code"):
        _set_cell(ws, meta["ba_code"], metadata["ba_code"])
    if metadata.get("annex_serial"):
        _set_cell(ws, meta["annex_serial"], metadata["annex_serial"])

    end_clear_row = start_row + max_rows - 1
    value_cols = list(cols.values())
    for row in range(start_row, end_clear_row + 1):
        for col_letter in value_cols:
            ws[f"{col_letter}{row}"].value = None

    cargo_default = defaults["cargo"].get(output_language, defaults["cargo"]["en"])
    flag_fields = [
        "loaded",
        "stackable",
        "rotatable",
        "weapons",
        "conditioned",
        "dangerous_goods",
        "ammunition",
        "itar",
        "tbb",
    ]
    included = [ln for ln in lines if ln.get("include", True)]
    for i, line in enumerate(included[:max_rows]):
        row = start_row + i
        if row != style_row:
            _copy_row_style(ws, style_row, row)
            _copy_formulas(ws, style_row, row, formula_cols)
        flags = line.get("appendix_flags") or {}
        _set_cell(ws, f"{cols['line_number']}{row}", i + 1)
        _set_cell(ws, f"{cols['cargo']}{row}", cargo_default)
        product = line.get("product_type") or "Item"
        type_label = product.replace("_", " ").title()
        _set_cell(ws, f"{cols['type']}{row}", type_label)
        _set_cell(ws, f"{cols['specifications']}{row}", line.get("output_description") or line.get("description"))
        _set_cell(ws, f"{cols['quantity']}{row}", line.get("quantity"))
        _set_cell(ws, f"{cols['registration']}{row}", f"{job_ref}:{i + 1}")
        _set_cell(ws, f"{cols['length_cm']}{row}", line.get("length_cm"))
        _set_cell(ws, f"{cols['width_cm']}{row}", line.get("width_cm"))
        _set_cell(ws, f"{cols['height_cm']}{row}", line.get("height_cm"))
        _set_cell(ws, f"{cols['weight_each_kg']}{row}", line.get("weight_each_kg"))
        for flag in flag_fields:
            _set_cell(ws, f"{cols[flag]}{row}", _yn(flags.get(flag), defaults.get(flag, "N")))
        if flags.get("temperature_c") and "temperature_c" in cols:
            _set_cell(ws, f"{cols['temperature_c']}{row}", flags.get("temperature_c"))
        if flags.get("tbb_category") and "tbb_category" in cols:
            _set_cell(ws, f"{cols['tbb_category']}{row}", flags.get("tbb_category"))

    if dangerous_goods:
        _fill_sheet_d(wb, mapping, metadata, dangerous_goods)

    # Alleen de relevante tabbladen A1 en D behouden; overige zijn overbodig.
    keep = {mapping["sheet"], (mapping.get("sheet_d") or {}).get("sheet")}
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep:
            del wb[sheet_name]

    fd, temp_name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    out_path = Path(temp_name)
    try:
        out_path.chmod(0o600)
    except OSError:
        pass
    wb.save(out_path)
    return out_path
