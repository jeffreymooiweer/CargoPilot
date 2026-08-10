import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.languages import normalise, pick
from app.services.dg.database import is_transport_forbidden
from app.services.dg.naming import resolve_for_profile
from app.services.documents.registry import condition_met, get_document, resolve_sections

TEXTS = {
    "generated_with": {
        "nl": "Gegenereerd met CargoPilot op",
        "en": "Generated with CargoPilot on",
        "de": "Erstellt mit CargoPilot am", "fr": 'Généré avec CargoPilot le'},
    "status": {"nl": "Documentstatus", "en": "Document status", "de": "Dokumentstatus", "fr": 'Statut du document'},
    "goods": {"nl": "Goederenregels", "en": "Cargo lines", "de": "Güterzeilen", "fr": 'Lignes de marchandises'},
    "dg_table": {"nl": "Gevaarlijke stoffen", "en": "Dangerous goods", "de": "Gefahrgut", "fr": 'Marchandises dangereuses'},
    "not_prefilled": {
        "nl": "niet vooraf ingevuld — handtekening/bevestiging vereist",
        "en": "not pre-filled — signature/confirmation required",
        "de": "nicht vorausgefüllt — Unterschrift/Bestätigung erforderlich", "fr": 'non prérempli — signature ou confirmation requise'},
    "carrier_provided": {
        "nl": "in te vullen door vervoerder/expediteur",
        "en": "to be provided by carrier/forwarder",
        "de": "vom Frachtführer/Spediteur auszufüllen", "fr": 'à fournir par le transporteur ou le commissionnaire'},
    "operational": {
        "nl": "in te vullen tijdens uitvoering",
        "en": "to be filled in during execution",
        "de": "während der Durchführung auszufüllen", "fr": "à compléter lors de l'exécution"},
    "confirmed": {
        "nl": "Bevestigd in CargoPilot; ondertekening op het document blijft vereist",
        "en": "Confirmed in CargoPilot; signature on the document is still required",
        "de": "In CargoPilot bestätigt; die Unterschrift auf dem Dokument bleibt erforderlich", "fr": 'Confirmé dans CargoPilot ; la signature sur le document reste requise'},
    "not_confirmed": {
        "nl": "NIET bevestigd",
        "en": "NOT confirmed",
        "de": "NICHT bestätigt", "fr": 'NON confirmé'},
    "totals": {"nl": "Totalen", "en": "Totals", "de": "Summen", "fr": 'Totaux'},
    "line_headers": {
        "nl": ["Nr", "Omschrijving", "Aantal", "Eenheid", "Gewicht (kg)", "Volume (m³)", "L×B×H (cm)"],
        "en": ["No", "Description", "Qty", "Unit", "Weight (kg)", "Volume (m³)", "L×W×H (cm)"],
        "de": ["Nr.", "Bezeichnung", "Menge", "Einheit", "Gewicht (kg)", "Volumen (m³)", "L×B×H (cm)"],
        "fr": ["N°", "Désignation", "Quantité", "Unité", "Poids (kg)", "Volume (m³)", "L×l×h (cm)"],
    },
    "dg_headers": {
        "nl": [
            "Positie",
            "UN-nummer",
            "Proper Shipping Name",
            "Technische naam",
            "Klasse",
            "Nevengevaar",
            "Verpakkingsgroep",
            "Packing instruction",
            "Aantal colli",
            "Verpakkingstype",
            "Hoeveelheid per verpakking",
            "Bruto massa per verpakking",
            "Marine pollutant",
            "Cargo Aircraft Only",
            "Aanvullende informatie",
        ],
        "en": [
            "Position",
            "UN number",
            "Proper Shipping Name",
            "Technical name",
            "Class",
            "Subsidiary risk",
            "Packing group",
            "Packing instruction",
            "Packages",
            "Package type",
            "Quantity per package",
            "Gross mass per package",
            "Marine pollutant",
            "Cargo Aircraft Only",
            "Additional information",
        ],
        "de": [
            "Position",
            "UN-Nummer",
            "Proper Shipping Name",
            "Technische Benennung",
            "Klasse",
            "Nebengefahr",
            "Verpackungsgruppe",
            "Verpackungsanweisung",
            "Anzahl Versandstücke",
            "Verpackungsart",
            "Menge je Verpackung",
            "Bruttomasse je Verpackung",
            "Meeresschadstoff",
            "Cargo Aircraft Only",
            "Zusätzliche Angaben",
        ],
        "fr": [
            "Position",
            "Numéro ONU",
            "Désignation officielle de transport",
            "Nom technique",
            "Classe",
            "Risque subsidiaire",
            "Groupe d'emballage",
            "Instruction d'emballage",
            "Nombre de colis",
            "Type d'emballage",
            "Quantité par emballage",
            "Masse brute par emballage",
            "Polluant marin",
            "Cargo Aircraft Only",
            "Informations complémentaires",
        ],
    },
    "dg_missing": {
        "nl": "Gevaarlijke-stoffenclassificatie onvolledig voor",
        "en": "Dangerous goods classification incomplete for",
        "de": "Gefahrgutklassifizierung unvollständig für", "fr": 'Classification des marchandises dangereuses incomplète pour'},
    "dg_forbidden": {
        "nl": "Niet ten vervoer toegelaten volgens ADR Tabel A (vervoer alleen onder ontheffing van de bevoegde autoriteit)",
        "en": "Not permitted for carriage per ADR Table A (carriage only under an exemption from the competent authority)",
        "de": "Nach ADR Tabelle A zur Beförderung nicht zugelassen (Beförderung nur mit Ausnahmegenehmigung der zuständigen Behörde)", "fr": "Non admis au transport selon le tableau A de l'ADR (transport uniquement sous dérogation de l'autorité compétente)"},
    "field_required": {
        "nl": "Verplicht veld ontbreekt",
        "en": "Required field missing",
        "de": "Pflichtfeld fehlt", "fr": 'Champ obligatoire manquant'},
    "dg_name_language": {
        "nl": "Vervoersnaam op dit document in het Engels gezet, zoals "
              "IMDG 5.4.1.4.1 / IATA DGR 8.1.2.1 voorschrijven",
        "en": "Proper shipping name set to English on this document, as "
              "IMDG 5.4.1.4.1 / IATA DGR 8.1.2.1 require",
        "de": "Offizielle Benennung auf diesem Dokument auf Englisch gesetzt, "
              "wie IMDG 5.4.1.4.1 / IATA DGR 8.1.2.1 es verlangen", "fr": "Désignation officielle de transport mise en anglais sur ce document, comme l'exigent le 5.4.1.4.1 de l'IMDG et le 8.1.2.1 de l'IATA DGR"},
    "field_format": {
        "nl": "Veld heeft niet de vereiste vorm",
        "en": "Field does not have the required format",
        "de": "Feld hat nicht die vorgeschriebene Form", "fr": "Le champ n'a pas le format requis"},
    "no_dg_lines": {
        "nl": "Dit document vereist gevaarlijke-stoffenregels, maar er zijn geen DG-posities.",
        "en": "This document requires dangerous goods lines, but no DG positions exist.",
        "de": "Dieses Dokument verlangt Gefahrgutzeilen, es sind aber keine Gefahrgutpositionen vorhanden.", "fr": "Ce document exige des lignes de marchandises dangereuses, mais aucune position n'a été saisie."},
    "vgm_mismatch": {
        "nl": "VGM wijkt af van de som van de componenten (methode 2)",
        "en": "VGM differs from the sum of the components (method 2)",
        "de": "VGM weicht von der Summe der Bestandteile ab (Methode 2)", "fr": 'La masse brute vérifiée diffère de la somme des composants (méthode 2)'},
    "fixed_texts": {
        "nl": "Vaste teksten en verklaringen (officieel formulier)",
        "en": "Fixed texts and declarations (official form)",
        "de": "Feste Texte und Erklärungen (amtliches Formular)", "fr": 'Textes fixes et déclarations (formulaire officiel)'},
    "legal_reference": {"nl": "Regelgeving", "en": "Regulations", "de": "Vorschriften", "fr": 'Réglementation'},
    "adr_points_incomplete": {
        "nl": "puntentelling onvolledig — controleer vervoerscategorie en hoeveelheid",
        "en": "points calculation incomplete — check transport category and quantity",
        "de": "Punkteberechnung unvollständig — Beförderungskategorie und Menge prüfen", "fr": 'calcul des points incomplet — vérifiez la catégorie de transport et la quantité'},
    "adr_exemption_lost": {
        "nl": "vrijstelling vervalt ({detail}) — de volledige ADR-eisen gelden",
        "en": "exemption does not apply ({detail}) — the full ADR requirements apply",
        "de": "Freistellung entfällt ({detail}) — es gelten die vollen ADR-Anforderungen", "fr": "l'exemption ne s'applique pas ({detail}) — l'ensemble des prescriptions de l'ADR s'applique"},
    "dg_description": {
        "nl": "Omschrijving vervoersdocument",
        "en": "Transport document description",
        "de": "Angabe im Beförderungspapier", "fr": 'Description du document de transport'},
    "disclaimer": {
        "nl": (
            "Dit document is automatisch gegenereerd met CargoPilot en is een concept: het moet vóór gebruik "
            "volledig worden gecontroleerd, aangevuld en ondertekend door een daartoe bevoegde persoon. "
            "CargoPilot en de maker(s) aanvaarden geen enkele aansprakelijkheid; de software wordt geleverd "
            "\"AS IS\" onder de Apache License 2.0 met Commons Clause (zie DISCLAIMER.md en LICENSE)."
        ),
        "en": (
            "This document was generated automatically with CargoPilot and is a draft: before use it must be "
            "fully verified, completed and signed by a duly authorised person. CargoPilot and its author(s) "
            "accept no liability whatsoever; the software is provided \"AS IS\" under the Apache License 2.0 "
            "with Commons Clause (see DISCLAIMER.md and LICENSE)."
        ),
        "de": (
            "Dieses Dokument wurde automatisch mit CargoPilot erstellt und ist ein Entwurf: Vor der "
            "Verwendung muss es von einer dazu befugten Person vollständig geprüft, ergänzt und "
            "unterschrieben werden. CargoPilot und seine Urheber übernehmen keinerlei Haftung; die "
            "Software wird \"AS IS\" unter der Apache License 2.0 mit Commons Clause bereitgestellt "
            "(siehe DISCLAIMER.md und LICENSE)."
        ), "fr": "Ce document a été généré automatiquement avec CargoPilot et constitue un projet : avant utilisation, il doit être intégralement vérifié, complété et signé par une personne dûment habilitée. CargoPilot et son ou ses auteurs déclinent toute responsabilité ; le logiciel est fourni « EN L'ÉTAT » sous licence Apache 2.0 avec Commons Clause (voir DISCLAIMER.md et LICENSE)."},
    "iata_dg_headers": {
        "nl": [
            "UN- of ID-nr.",
            "Proper Shipping Name (technische naam)",
            "Klasse of divisie (nevengevaar)",
            "Verpakkingsgroep",
            "Hoeveelheid en soort verpakking",
            "Packing Inst.",
            "Authorization",
        ],
        "en": [
            "UN or ID No.",
            "Proper Shipping Name (technical name)",
            "Class or Division (Subsidiary Hazard)",
            "Packing Group",
            "Quantity and Type of Packing",
            "Packing Inst.",
            "Authorization",
        ],
        "de": [
            "UN- oder ID-Nr.",
            "Proper Shipping Name (technische Benennung)",
            "Klasse oder Unterklasse (Nebengefahr)",
            "Verpackungsgruppe",
            "Menge und Verpackungsart",
            "Packing Inst.",
            "Authorization",
        ],
        "fr": [
            "N° ONU ou ID",
            "Désignation officielle de transport (nom technique)",
            "Classe ou division (risque subsidiaire)",
            "Groupe d'emballage",
            "Quantité et type d'emballage",
            "Packing Inst.",
            "Authorization",
        ],
    },
    "adr_dg_headers": {
        "nl": [
            "Omschrijving conform 5.4.1.1.1",
            "Aantal colli",
            "Verpakkingstype",
            "Hoeveelheid per verpakking",
            "Bruto massa per verpakking",
            "Aanvullende informatie",
        ],
        "en": [
            "Description per 5.4.1.1.1",
            "Packages",
            "Package type",
            "Quantity per package",
            "Gross mass per package",
            "Additional information",
        ],
        "de": [
            "Angabe nach 5.4.1.1.1",
            "Anzahl Versandstücke",
            "Verpackungsart",
            "Menge je Verpackung",
            "Bruttomasse je Verpackung",
            "Zusätzliche Angaben",
        ],
        "fr": [
            "Description selon le 5.4.1.1.1",
            "Nombre de colis",
            "Type d'emballage",
            "Quantité par emballage",
            "Masse brute par emballage",
            "Informations complémentaires",
        ],
    },
    "imdg_dg_headers": {
        "nl": [
            "UN-nummer",
            "Proper Shipping Name (technische naam)",
            "Klasse (nevengevaar)",
            "Verpakkingsgroep",
            "Marine pollutant",
            "Vlampunt",
            "EmS",
            "Aantal en soort colli",
            "Hoeveelheid per verpakking",
            "Bruto massa per verpakking",
        ],
        "en": [
            "UN number",
            "Proper Shipping Name (technical name)",
            "Class (subsidiary risk)",
            "Packing group",
            "Marine pollutant",
            "Flashpoint",
            "EmS",
            "Number and kind of packages",
            "Quantity per package",
            "Gross mass per package",
        ],
        "fr": [
            "Numéro ONU",
            "Désignation officielle de transport (nom technique)",
            "Classe (risque subsidiaire)",
            "Groupe d'emballage",
            "Polluant marin",
            "Point d'éclair",
            "EmS",
            "Nombre et type de colis",
            "Quantité par emballage",
            "Masse brute par emballage",
        ],
        "de": [
            "UN-Nummer",
            "Proper Shipping Name (technische Benennung)",
            "Klasse (Nebengefahr)",
            "Verpackungsgruppe",
            "Meeresschadstoff",
            "Flammpunkt",
            "EmS",
            "Anzahl und Art der Versandstücke",
            "Menge je Verpackung",
            "Bruttomasse je Verpackung",
        ],
    },
}

DG_PRODUCT_FIELDS = [
    "un_number",
    "proper_shipping_name",
    "technical_name",
    "class",
    "subsidiary_risks",
    "packing_group",
    "packing_instruction",
    "quantity_packages",
    "type_of_package",
    "net_mass_liters_per_package",
    "gross_mass_per_package",
    "marine_pollutant",
    "cargo_aircraft_only",
    "additional_information",
]

DG_BASE_REQUIRED = ["un_number", "proper_shipping_name", "class"]
DG_PROFILE_REQUIRED = {
    "ADR": DG_BASE_REQUIRED,
    "RID": DG_BASE_REQUIRED,
    "ADN": DG_BASE_REQUIRED,
    "IMDG": DG_BASE_REQUIRED + ["quantity_packages", "type_of_package"],
    "IATA_DGR": DG_BASE_REQUIRED
    + ["packing_instruction", "quantity_packages", "type_of_package", "net_mass_liters_per_package"],
}


def _lang(language: str) -> str:
    return normalise(language)


def _label(item: dict[str, Any], lang: str) -> str:
    return pick(item.get("label"), lang, item.get("key", "")) or item.get("key", "")


def _localised(value: Any, lang: str) -> str:
    """A {nl, en, de} block from the registry in the requested language."""
    if isinstance(value, dict):
        return str(pick(value, lang))
    return str(value or "")


def _text(key: str, lang: str) -> Any:
    # A fallback rather than a KeyError: a language still missing a single line
    # must not bring an export down.
    return pick(TEXTS[key], lang)


def _option_label(field: dict[str, Any], value: Any, lang: str) -> Any:
    for option in field.get("options", []):
        if option.get("value") == value:
            return _label(option, lang)
    return value


def validate_document(
    document: dict[str, Any],
    values: dict[str, Any],
    lines: list[dict[str, Any]],
    dangerous_goods: list[dict[str, Any]] | None,
    language: str = "nl",
) -> tuple[list[str], list[str]]:
    """Return (blocking errors, warnings) for a document export."""
    lang = _lang(language)
    errors: list[str] = []
    warnings: list[str] = []

    for section in resolve_sections(document):
        for field in section.get("fields", []):
            value = values.get(field["key"])
            empty = value is None or str(value).strip() == ""
            if field.get("status") == "USER_REQUIRED" and empty:
                errors.append(f"{_text('field_required', lang)}: {_label(field, lang)}")
                continue
            # A field that promises a shape has to have that shape. Until now
            # only emptiness was checked, so "72" or "7208 51" simply ended up on
            # an official waybill as the NHM code.
            pattern = field.get("pattern")
            if pattern and not empty and not re.fullmatch(pattern, str(value).strip()):
                errors.append(
                    f"{_text('field_format', lang)}: {_label(field, lang)}"
                    + (f" — {_localised(field.get('format_hint'), lang)}"
                       if field.get("format_hint") else "")
                )

    profile = document.get("dg_profile")
    if profile:
        entries = dangerous_goods or []
        if document.get("dg_only") and not entries:
            errors.append(_text("no_dg_lines", lang))
        required_fields = DG_PROFILE_REQUIRED.get(profile, DG_BASE_REQUIRED)
        for entry in entries:
            for product in entry.get("products", []):
                # Substances ADR does not admit for carriage block the export.
                if is_transport_forbidden(str(product.get("un_number") or "")):
                    errors.append(
                        f"{_text('dg_forbidden', lang)}: {_un_prefixed(product.get('un_number'))}"
                    )
                # Sea and air prescribe the language of the name: IMDG 5.4.1.4.1
                # permits English, French or Spanish and IATA DGR 8.1.2.1 only
                # English. Whoever first drew up a German road document keeps the
                # German name standing in the field.
                #
                # That is no reason to refuse the export: the language belongs to
                # the document and not to the consignment, and CargoPilot knows
                # which name has to be here. It puts it there itself and reports
                # it — blocking would only make the user retype what the app
                # already knew.
                english, replaced = resolve_for_profile(product, profile)
                if replaced:
                    warnings.append(
                        f"{_text('dg_name_language', lang)}: "
                        f"{_un_prefixed(product.get('un_number'))} — "
                        f"{replaced} → {english}"
                    )
                missing = [f for f in required_fields if not str(product.get(f) or "").strip()]
                if missing:
                    position = entry.get("vehicle") or entry.get("line_id") or "?"
                    errors.append(
                        f"{_text('dg_missing', lang)} '{position}': {', '.join(missing)}"
                    )

        # The full compliance check runs at the export itself as well. The panel
        # in the wizard is an aid; the frontend must never be the only place
        # where this is enforced — a stale or never-refreshed screen result must
        # not produce a document.
        if entries:
            from app.services.dg.compliance import check_compliance

            outcome = check_compliance(entries, [profile], language)
            # If this export computes with a rule set that has run out and has
            # not been replaced, that belongs on the document and not only on the
            # screen — a document outlives the session it was made in.
            for finding in outcome.get("rule_set_warnings", []) or []:
                warnings.append(f"{finding['rule']}: {finding['message']}")
            for finding in outcome.get("imdg_segregation", []) + outcome.get(
                "adr_mixed_loading", []
            ) + outcome.get("iata_segregation", []):
                text = f"{finding.get('rule')}: {finding.get('message')}"
                if finding.get("severity") == "error":
                    errors.append(text)
                elif finding.get("severity") == "warning":
                    warnings.append(text)
            for q in outcome.get("q_values", []) or []:
                if q.get("status") == "exceeded":
                    # Q above 1 means the combination may not fly like that.
                    errors.append(
                        f"IATA 5.0.2.11: Q = {q.get('q_value')} (> 1)"
                        + (f" — {q.get('position')}" if q.get("position") else "")
                    )
                elif q.get("status") in {"incomplete", "not_checked"}:
                    warnings.append(str(q.get("note") or "Q incomplete"))
            # And if no Q was computed at all, that belongs on it too: a check
            # that did not run looked on the document like a check that passed.
            q_status = outcome.get("q_check_status")
            if q_status and q_status.get("status") in {"not_checked", "incomplete"}:
                warnings.append(f"IATA DGR 5.0.2.11: {q_status['message']}")
            adn = outcome.get("adn_exemption")
            if adn and profile == "ADN":
                if adn.get("status") == "incomplete":
                    warnings.append(
                        "ADN 1.1.3.6.1: " + _text("adr_points_incomplete", lang)
                    )
                elif adn.get("status") in {"not_exempt", "above_threshold"}:
                    over = ", ".join(
                        f"class {item['class']} {item['carried']} kg > {item['limit']} kg"
                        for item in adn.get("over_class_limit") or []
                    )
                    detail = over or (
                        f"{adn.get('total_gross_mass_kg')} kg > {adn.get('threshold')} kg"
                    )
                    warnings.append(
                        "ADN 1.1.3.6.1: " + _text("adr_exemption_lost", lang).format(detail=detail)
                    )
            points = outcome.get("adr_points")
            if points and profile in {"ADR", "RID", "ADN"}:
                status = points.get("status")
                if status == "incomplete":
                    warnings.append("ADR 1.1.3.6: " + _text("adr_points_incomplete", lang))
                elif status in {"above_threshold", "not_exempt"}:
                    # Not forbidden, but the 1.1.3.6 exemption lapses and with
                    # it the full requirements apply: training, ADR vehicle,
                    # orange plates, fire extinguishers. Whoever still saw
                    # "exemption possible" on screen has to read this here.
                    total = points.get("total_points")
                    threshold = points.get("threshold")
                    detail = (
                        f"{total} > {threshold}" if status == "above_threshold"
                        else ", ".join(points.get("category0_products") or [])
                    )
                    warnings.append(
                        "ADR 1.1.3.6: "
                        + _text("adr_exemption_lost", lang).format(detail=detail)
                    )
            # The tunnel code of column (15) is on the document per 5.4.1.1.1 (k)
            # and, until v1.50.0, evaluated nowhere. What the driver needs is the
            # code of the *whole load* (8.6.3.2), and that never appeared on the
            # sheet at all — only the per-substance codes it is derived from.
            tunnel = outcome.get("adr_tunnel")
            if tunnel and profile == "ADR" and tunnel.get("status") not in {None, "not_checked"}:
                warnings.append(f"ADR 8.6.3: {tunnel['message']}")

    if document["key"] == "vgm" and str(values.get("vgm_method")) == "method2":
        components = [
            values.get("cargo_mass_kg"),
            values.get("packaging_mass_kg"),
            values.get("pallets_mass_kg"),
            values.get("securing_mass_kg"),
            values.get("container_tare_kg"),
        ]
        try:
            total = sum(float(v) for v in components if v not in (None, ""))
            vgm = float(values.get("vgm_kg") or 0)
            if vgm and abs(total - vgm) > max(0.005 * vgm, 1.0):
                warnings.append(
                    f"{_text('vgm_mismatch', lang)}: {vgm} kg ≠ {round(total, 2)} kg"
                )
        except (TypeError, ValueError):
            pass

    return errors, warnings


def _un_prefixed(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text if text.upper().startswith(("UN", "ID")) else f"UN {text}"


def _dg_description(product: dict[str, Any], profile: str, values: dict[str, Any]) -> str:
    """Officiële omschrijvingsregel per ADR/RID/ADN 5.4.1.1.1, bijv. 'UN 1203, BENZINE, 3, II, (D/E)'."""
    psn = resolve_for_profile(product, profile)[0].upper()
    technical = str(product.get("technical_name") or "").strip()
    if technical:
        psn = f"{psn} ({technical})"
    hazard = str(product.get("class") or "").strip()
    subsidiary = str(product.get("subsidiary_risks") or "").strip()
    if subsidiary:
        hazard = f"{hazard} ({subsidiary})"
    parts = [_un_prefixed(product.get("un_number")), psn, hazard, str(product.get("packing_group") or "").strip()]
    # ADR only: the tunnel restriction code is a road construct (Table A column
    # 15, 8.6, 5.4.1.1.1 (k)). It does not belong on a CIM or an ADN document.
    if profile == "ADR":
        # Tunnel code from the UN entry; a manual value takes precedence.
        tunnel = str(values.get("tunnel_restriction") or product.get("tunnel_code") or "").strip()
        if tunnel:
            parts.append(f"({tunnel.strip('()')})")
    return ", ".join(p for p in parts if p)


def _dg_rows(profile: str, entry: dict[str, Any], product: dict[str, Any], values: dict[str, Any], lang: str):
    """Row values for the DG table, in the column order of the form concerned."""
    if profile == "IATA_DGR":
        quantity_parts = [
            str(product.get("quantity_packages") or "").strip(),
            str(product.get("type_of_package") or "").strip(),
        ]
        quantity = " × ".join(p for p in quantity_parts if p)
        per_package = str(product.get("net_mass_liters_per_package") or "").strip()
        if per_package:
            quantity = f"{quantity}, {per_package}" if quantity else per_package
        psn = resolve_for_profile(product, profile)[0]
        technical = str(product.get("technical_name") or "").strip()
        if technical:
            psn = f"{psn} ({technical})"
        hazard = str(product.get("class") or "")
        subsidiary = str(product.get("subsidiary_risks") or "").strip()
        if subsidiary:
            hazard = f"{hazard} ({subsidiary})"
        return [
            _un_prefixed(product.get("un_number")),
            psn,
            hazard,
            product.get("packing_group", ""),
            quantity,
            product.get("packing_instruction", ""),
            product.get("authorization", "") or product.get("eq_lq_points", ""),
        ]
    if profile in {"ADR", "RID", "ADN"}:
        return [
            _dg_description(product, profile, values),
            product.get("quantity_packages", ""),
            product.get("type_of_package", ""),
            product.get("net_mass_liters_per_package", ""),
            product.get("gross_mass_per_package", ""),
            product.get("additional_information", ""),
        ]
    if profile == "IMDG":
        psn = resolve_for_profile(product, profile)[0]
        technical = str(product.get("technical_name") or "").strip()
        if technical:
            psn = f"{psn} ({technical})"
        hazard = str(product.get("class") or "")
        subsidiary = str(product.get("subsidiary_risks") or "").strip()
        if subsidiary:
            hazard = f"{hazard} ({subsidiary})"
        packages_parts = [
            str(product.get("quantity_packages") or "").strip(),
            str(product.get("type_of_package") or "").strip(),
        ]
        return [
            _un_prefixed(product.get("un_number")),
            psn,
            hazard,
            product.get("packing_group", ""),
            product.get("marine_pollutant", ""),
            product.get("flashpoint", ""),
            product.get("ems_code", ""),
            " × ".join(p for p in packages_parts if p),
            product.get("net_mass_liters_per_package", ""),
            product.get("gross_mass_per_package", ""),
        ]
    return [entry.get("vehicle") or entry.get("line_id")] + [
        product.get(field, "") for field in DG_PRODUCT_FIELDS
    ]


def _dg_headers(profile: str, lang: str) -> list[str]:
    if profile == "IATA_DGR":
        return _text("iata_dg_headers", lang)
    if profile in {"ADR", "RID", "ADN"}:
        return _text("adr_dg_headers", lang)
    if profile == "IMDG":
        return _text("imdg_dg_headers", lang)
    return _text("dg_headers", lang)


def _dims(line: dict[str, Any]) -> str:
    parts = [line.get("length_cm"), line.get("width_cm"), line.get("height_cm")]
    if all(p in (None, "") for p in parts):
        return ""
    return " × ".join(str(p) if p not in (None, "") else "—" for p in parts)


def export_document(
    document_key: str,
    values: dict[str, Any],
    lines: list[dict[str, Any]],
    dangerous_goods: list[dict[str, Any]] | None,
    language: str = "nl",
) -> Path:
    document = get_document(document_key)
    if document is None:
        raise ValueError(f"Unknown document: {document_key}")
    lang = _lang(language)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = document.get("short_label", {}).get(lang, document_key)[:31]

    title_font = Font(bold=True, size=15)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E2EC")
    note_font = Font(italic=True, size=9, color="666666")
    status_font = Font(italic=True, size=10, color="1E3A5F")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 42
    for col in range(2, 16):
        ws.column_dimensions[get_column_letter(col)].width = 22

    row = 1
    ws.cell(row, 1, _label(document, lang)).font = title_font
    row += 1
    issue = pick(document.get("issue_status"), lang)
    if issue:
        cell = ws.cell(row, 1, f"{_text('status', lang)}: {issue}")
        cell.font = status_font
        row += 1
    ws.cell(row, 1, f"{_text('generated_with', lang)} {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = note_font
    row += 2

    for section in resolve_sections(document):
        fields = section.get("fields", [])
        visible = [
            f
            for f in fields
            if f.get("status") != "CONDITIONAL"
            or not f.get("condition")
            or condition_met(f.get("condition"), values)
            or str(values.get(f["key"], "")).strip() != ""
        ]
        filled_or_relevant = [
            f
            for f in visible
            if str(values.get(f["key"], "")).strip() != ""
            or f.get("status") in {"USER_REQUIRED", "CARRIER_PROVIDED", "OPERATIONAL", "SIGNATURE_REQUIRED"}
        ]
        if not filled_or_relevant:
            continue
        cell = ws.cell(row, 1, _label(section, lang))
        cell.font = section_font
        cell.fill = section_fill
        ws.cell(row, 2).fill = section_fill
        row += 1
        for field in filled_or_relevant:
            value = values.get(field["key"], "")
            status = field.get("status")
            if field.get("type") == "select" and value not in (None, ""):
                value = _option_label(field, value, lang)
            if status == "SIGNATURE_REQUIRED":
                if field.get("type") == "checkbox":
                    value = (
                        _text("confirmed", lang)
                        if str(value).lower() in {"true", "1", "yes", "ja"}
                        else _text("not_confirmed", lang)
                    )
                else:
                    value = f"[{_text('not_prefilled', lang)}]"
            elif str(value).strip() == "":
                if status == "CARRIER_PROVIDED":
                    value = f"[{_text('carrier_provided', lang)}]"
                elif status == "OPERATIONAL":
                    value = f"[{_text('operational', lang)}]"
                else:
                    value = ""
            label_cell = ws.cell(row, 1, _label(field, lang))
            label_cell.alignment = wrap
            label_cell.border = border
            value_cell = ws.cell(row, 2, value if value != "" else None)
            value_cell.alignment = wrap
            value_cell.border = border
            if isinstance(value, str) and value.startswith("["):
                value_cell.font = note_font
            row += 1
        row += 1

    included = [ln for ln in lines if ln.get("include", True)]
    if included:
        cell = ws.cell(row, 1, _text("goods", lang))
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        headers = _text("line_headers", lang)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        total_weight = 0.0
        total_volume = 0.0
        for i, line in enumerate(included, start=1):
            weight = line.get("weight_total_kg")
            volume = line.get("transport_volume_m3")
            total_weight += weight or 0
            total_volume += volume or 0
            cells = [
                i,
                line.get("output_description") or line.get("description"),
                line.get("quantity"),
                line.get("unit"),
                weight,
                volume,
                _dims(line),
            ]
            for col, value in enumerate(cells, start=1):
                cell = ws.cell(row, col, value)
                cell.border = border
                if col == 2:
                    cell.alignment = wrap
            row += 1
        cell = ws.cell(row, 1, _text("totals", lang))
        cell.font = header_font
        ws.cell(row, 5, round(total_weight, 2)).font = header_font
        ws.cell(row, 6, round(total_volume, 3)).font = header_font
        row += 2

    if document.get("dg_profile") and dangerous_goods:
        profile = document["dg_profile"]
        cell = ws.cell(row, 1, f"{_text('dg_table', lang)} ({profile})")
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        headers = _dg_headers(profile, lang)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = wrap
        row += 1
        for entry in dangerous_goods:
            for product in entry.get("products", []):
                for col, value in enumerate(_dg_rows(profile, entry, product, values, lang), start=1):
                    cell = ws.cell(row, col, value)
                    cell.border = border
                    cell.alignment = wrap
                row += 1
        row += 1

    fixed_texts = document.get("fixed_texts") or []
    if fixed_texts:
        cell = ws.cell(row, 1, _text("fixed_texts", lang))
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        for item in fixed_texts:
            text = pick(item, lang)
            cell = ws.cell(row, 1, text)
            cell.alignment = wrap
            cell.border = border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.row_dimensions[row].height = max(28, 13 * (len(text) // 110 + 1))
            row += 1
        row += 1

    legal = pick(document.get("legal_reference"), lang)
    if legal:
        cell = ws.cell(row, 1, f"{_text('legal_reference', lang)}: {legal}")
        cell.font = note_font
        cell.alignment = wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

    note = pick(document.get("signature_note"), lang)
    if note:
        cell = ws.cell(row, 1, note)
        cell.font = note_font
        cell.alignment = wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

    row += 1
    cell = ws.cell(row, 1, _text("disclaimer", lang))
    cell.font = note_font
    cell.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 40

    fd, temp_name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    out_path = Path(temp_name)
    try:
        out_path.chmod(0o600)
    except OSError:
        pass
    wb.save(out_path)
    return out_path
