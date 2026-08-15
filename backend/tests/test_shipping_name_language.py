"""The proper shipping name cannot simply be translated along with the screen.

The ADR table carries two official names per UN number — `name_en` and `name_de`
— and the app always took the English one. A German consignor got "GASOLINE" on
their CMR while "BENZIN ODER OTTOKRAFTSTOFF" was sitting next to it in Table A.

Putting that right cannot be done with one line, because the modes differ:

* ADR 5.4.1.4.1 (and along the same lines RID/ADN) permits a German name;
* IMDG 5.4.1.4.1 requires English, French or Spanish;
* IATA DGR 8.1.2.1 requires English.

"BENZIN" on a Shipper's Declaration is not a matter of translation taste but a
refused consignment. These tests record that boundary — including the case where
somebody first draws up a German road document and then adds a sea leg.
"""

import pytest

from app.services.dg.database import get_un_entries, offline_lookup
from app.services.dg.naming import (
    is_german_name,
    proper_shipping_name,
    requires_english_name,
    resolve_for_profile,
)

# UN 1203 is called something different in German from English in Table A; that
# makes it suitable for reading the difference off.
BENZINE = "1203"

#: What column (2) of the 2025 volume prints for it. The 2023 export kept
#: only the middle alternative — "Gasoline" — and since v1.89.0 the English
#: column is read from the book, which sets all three.
PETROL = "MOTOR SPIRIT OR GASOLINE OR PETROL"


def entry(un: str) -> dict:
    entries = get_un_entries(un)
    assert entries, f"UN {un} ontbreekt in de ADR-tabel"
    return entries[0]


def test_the_adr_table_actually_carries_german_names():
    """Without this column the rest of this file has no ground to stand on."""
    from app.services.dg.naming import english_name_of

    assert entry(BENZINE)["name_de"].upper().startswith("BENZIN")
    assert english_name_of(entry(BENZINE)) == PETROL
    # The row itself still carries the export's single alternative; the column
    # the document uses is the one read from the 2025 volume.
    assert entry(BENZINE)["name_en"].upper() == "GASOLINE"


@pytest.mark.parametrize("profiles", [["ADR"], ["RID"], ["ADN"], ["ADR", "RID"], []])
def test_a_land_document_in_german_gets_the_german_name(profiles):
    assert proper_shipping_name(entry(BENZINE), "de", profiles) == "BENZIN ODER OTTOKRAFTSTOFF"


@pytest.mark.parametrize("profiles", [["IMDG"], ["IATA_DGR"], ["ADR", "IMDG"], ["ADR", "IATA_DGR"]])
def test_sea_and_air_keep_english_whatever_the_screen_says(profiles):
    """For a multimodal consignment English satisfies all three regimes; German
    only one of them."""
    assert proper_shipping_name(entry(BENZINE), "de", profiles) == PETROL


def test_an_english_reader_keeps_the_english_name():
    assert proper_shipping_name(entry(BENZINE), "en", ["ADR"]) == PETROL


def test_a_french_road_document_carries_the_french_name_alone():
    """The ADR is authentic in French too, and table A prints the column: a
    French reader gets the name the French edition gives, which 5.4.1.4.1 lets
    stand on its own."""
    assert proper_shipping_name(entry(BENZINE), "fr", ["ADR"]) == "ESSENCE"


def test_sea_and_air_take_the_french_name_off_again():
    assert proper_shipping_name(entry(BENZINE), "fr", ["IMDG"]) == PETROL


# --- Dutch: not one of the three ------------------------------------------
#
# ADR 5.4.1.4.1 asks for an official language of the forwarding country and,
# where that is not English, French or German, one of those three *in addition*.
# So Dutch is the one language that cannot stand alone on the document, and the
# field carries both names.


@pytest.mark.parametrize("language", ["nl", "", "xx"])
def test_a_dutch_road_document_carries_the_dutch_name_and_the_english_beside_it(language):
    """An unset or unknown language is Dutch — that is the app's default — and so
    it lands on the same rule rather than quietly on the English one."""
    assert (proper_shipping_name(entry(BENZINE), language, ["ADR"])
            == "BENZINE OF MOTORBRANDSTOF (" + PETROL + ")")


@pytest.mark.parametrize("un,expected", [
    ("1789", "ZOUTZUUR (HYDROCHLORIC ACID)"),
    ("1090", "ACETON (ACETONE)"),
])
def test_the_dutch_name_comes_from_the_adr_and_not_from_a_translation(un, expected):
    """These names are read out of table A of the Dutch ADR edition, not
    translated. ZOUTZUUR is what the book says; a translator would have produced
    "WATERSTOFCHLORIDE-OPLOSSING" or worse."""
    assert proper_shipping_name(entry(un), "nl", ["ADR"]) == expected


@pytest.mark.parametrize("profiles", [["IMDG"], ["IATA_DGR"], ["ADR", "IMDG"]])
def test_sea_and_air_get_the_english_name_alone_from_a_dutch_screen_too(profiles):
    """Two names in one field is what the road document wants; IMDG 5.4.1.4.1
    and IATA DGR 8.1.2.1 want one, in English."""
    assert proper_shipping_name(entry(BENZINE), "nl", profiles) == PETROL


def test_a_un_number_without_a_dutch_name_keeps_the_english_one_alone():
    """Not every entry the app knows is in the Dutch table A — the IMDG-only
    additions are not. An empty pair of brackets would be worse than nothing."""
    assert proper_shipping_name(
        {"un": "9999", "name_en": "SOMETHING", "name_de": ""}, "nl", ["ADR"]
    ) == "SOMETHING"


def test_a_name_that_reads_the_same_in_both_languages_is_not_doubled():
    """"TOLUEEN (TOLUEEN)" is not a requirement met, it is a printing error."""
    assert proper_shipping_name(
        {"un": "9999", "name_en": "TOLUENE", "name_de": "", "name_nl": "TOLUENE"},
        "nl", ["ADR"],
    ) == "TOLUENE"


def test_the_dutch_document_name_is_corrected_for_a_sea_leg():
    """Same trap as with German: a road document drawn up first keeps its name in
    the field, and on the IMO DGF only the English belongs."""
    product = {"un_number": "1203",
               "proper_shipping_name": "BENZINE OF MOTORBRANDSTOF (" + PETROL + ")"}
    assert resolve_for_profile(product, "IMDG") == (
        PETROL, "BENZINE OF MOTORBRANDSTOF (" + PETROL + ")")
    assert resolve_for_profile(dict(product), "ADR")[0] == (
        "BENZINE OF MOTORBRANDSTOF (" + PETROL + ")")


def test_an_entry_without_a_german_name_falls_back_instead_of_going_blank():
    # The IMDG-only entries from 42-24 carry no German name.
    assert proper_shipping_name({"name_en": "DISILANE", "name_de": ""}, "de", ["ADR"]) == "DISILANE"


def test_requires_english_name_is_case_and_whitespace_proof():
    assert requires_english_name(["imdg"])
    assert requires_english_name([" IATA_DGR "])
    assert not requires_english_name(["ADR", "RID"])
    assert not requires_english_name(None)


# --- The dangerous case: road first, sea afterwards ------------------------
#
# The language of the name belongs to the document, not to the consignment. One
# consignment produces a CMR with the German name and an IMO DGF with the
# English, from the same data. Refusing the export and making the user retype
# "GASOLINE" would have them do what the app already knows.


def german_goods():
    return [{
        "line_id": "L1",
        "products": [{
            "un_number": "1203",
            "proper_shipping_name": "BENZIN ODER OTTOKRAFTSTOFF",
            "class": "3",
            "packing_group": "II",
            "quantity_packages": "4",
            "type_of_package": "1A1",
            "net_mass_liters_per_package": "20",
        }],
    }]


@pytest.mark.parametrize("profile,expected", [
    ("ADR", "BENZIN ODER OTTOKRAFTSTOFF"),
    ("RID", "BENZIN ODER OTTOKRAFTSTOFF"),
    ("ADN", "BENZIN ODER OTTOKRAFTSTOFF"),
    ("IMDG", PETROL),
    ("IATA_DGR", PETROL),
])
def test_the_document_gets_the_name_its_own_rulebook_wants(profile, expected):
    name, _ = resolve_for_profile(german_goods()[0]["products"][0], profile)
    assert name == expected


def test_a_sea_document_is_not_refused_but_corrected():
    from app.services.documents.exporter import validate_document
    from app.services.documents.registry import get_document
    from tests.test_documents import BASE_VALUES, LINES

    errors, warnings = validate_document(
        get_document("imo_dgd"), dict(BASE_VALUES), LINES, german_goods(), "de"
    )
    assert [e for e in errors if "5.4.1.4.1" in e] == [], "export mag hier niet blokkeren"
    said = [w for w in warnings if "5.4.1.4.1" in w]
    assert said, warnings
    # The message says what happened, not what the user still has to do.
    assert "BENZIN ODER OTTOKRAFTSTOFF" in said[0] and PETROL in said[0]


def test_the_road_document_keeps_the_german_name_and_says_nothing():
    from app.services.documents.exporter import validate_document
    from app.services.documents.registry import get_document
    from tests.test_documents import BASE_VALUES, LINES

    errors, warnings = validate_document(
        get_document("cmr"), dict(BASE_VALUES), LINES, german_goods(), "de"
    )
    assert [e for e in errors if "5.4.1.4.1" in e] == []
    assert [w for w in warnings if "5.4.1.4.1" in w] == []


def test_the_exported_sea_document_actually_carries_the_english_name():
    """The warning is not enough — it has to actually be on the sheet."""
    import openpyxl

    from app.services.documents.exporter import export_document
    from tests.test_documents import BASE_VALUES, LINES

    path = export_document(
        "imo_dgd", dict(BASE_VALUES), LINES, german_goods(), language="de"
    )
    text = "\n".join(
        str(cell)
        for row in openpyxl.load_workbook(path).active.iter_rows(values_only=True)
        for cell in row
        if cell
    )
    assert PETROL in text
    assert "BENZIN ODER OTTOKRAFTSTOFF" not in text


def test_the_exported_road_document_carries_the_german_name():
    import openpyxl

    from app.services.documents.exporter import export_document
    from tests.test_documents import BASE_VALUES, LINES

    path = export_document("cmr", dict(BASE_VALUES), LINES, german_goods(), language="de")
    text = "\n".join(
        str(cell)
        for row in openpyxl.load_workbook(path).active.iter_rows(values_only=True)
        for cell in row
        if cell
    )
    assert "BENZIN ODER OTTOKRAFTSTOFF" in text


def test_the_description_line_follows_the_document_too():
    """The 5.4.1.1.1 line is the text that goes into the goods column verbatim."""
    from app.services.dg.autofill import description_line

    product = german_goods()[0]["products"][0]
    assert "BENZIN ODER OTTOKRAFTSTOFF" in description_line(product, "ADR")
    assert PETROL in description_line(product, "IMDG")
    assert "BENZIN" not in description_line(product, "IMDG")


def test_wording_the_user_wrote_themselves_is_left_alone():
    """A technical name with an n.o.s. entry or an addition of the user's own is
    something we cannot assess, and certainly must not replace silently."""
    own = {"un_number": "1203", "proper_shipping_name": "BENZIN, ENTHÄLT ETHANOL"}
    name, replaced = resolve_for_profile(own, "IMDG")
    assert name == "BENZIN, ENTHÄLT ETHANOL"
    assert replaced == ""


def test_an_english_name_is_not_touched_and_not_reported():
    name, replaced = resolve_for_profile(
        {"un_number": "1203", "proper_shipping_name": PETROL}, "IMDG"
    )
    assert (name, replaced) == (PETROL, "")


def test_an_empty_name_stays_empty_rather_than_being_invented():
    """A missing name is a separate fault; that is already reported as a missing
    field and must not be quietly filled in here."""
    assert resolve_for_profile({"un_number": "1203"}, "IMDG") == ("", "")


def test_an_english_name_never_trips_the_check():
    assert not is_german_name(entry(BENZINE), PETROL)


def test_a_name_that_reads_the_same_in_both_languages_is_not_flagged():
    """Many entries read the same in both languages; there is nothing to report
    about those and a warning would be nothing but noise."""
    assert not is_german_name({"name_en": "TOLUENE", "name_de": "TOLUENE"}, "TOLUENE")


def test_a_technical_name_in_brackets_is_not_flagged():
    """With an n.o.s. entry the consignor adds a technical name themselves; that
    is not a language fault."""
    nos = entry("3082")
    assert not is_german_name(nos, f"{nos['name_en']} (ALLYLALCOHOL)")


# --- The lookup produces the same name as the export ----------------------


@pytest.mark.parametrize("profiles,expected", [
    (["ADR"], "BENZIN ODER OTTOKRAFTSTOFF"),
    (["IMDG"], PETROL),
])
def test_the_lookup_suggests_the_name_that_the_document_will_carry(profiles, expected):
    """The suggestion the user clicks *is* the text that ends up on the document;
    those two must not diverge."""
    result = offline_lookup(BENZINE, "de", profiles)
    assert result["proper_shipping_name"] == expected


# --- Entries the table has no English name for ----------------------------
#
# Fourteen of them, plus one truncated — in the 2023 export. Reading column (2)
# out of the 2025 volume names every one of them, which is the whole argument
# for reading the book. The check stays: it is what makes a future edition's
# gap visible, and one entry still needs it.


@pytest.mark.parametrize("un", ["3245", "3374", "2807", "1327"])
def test_the_book_names_what_the_export_left_empty(un):
    """Fourteen entries carried an empty English name in the export — UN 3245
    genetically modified organisms, UN 3374 acetylene solvent free, UN 2807
    magnetized material among them — and a German name went on the document in
    place of an English one. The 2025 volume names them all."""
    from app.services.dg.naming import english_name_is_usable, english_name_of

    assert not str(entry(un).get("name_en") or "").strip()
    assert english_name_of(entry(un))
    assert english_name_is_usable(entry(un))


def test_the_book_names_the_entry_the_export_cut_off():
    """UN 1139 reads "Coating solution (" in the export — cut off mid-bracket.
    A name that ends on an opening bracket is not a name, and the volume has
    the whole of it."""
    from app.services.dg.naming import english_name_is_usable, english_name_of

    assert entry("1139")["name_en"] == "Coating solution ("
    assert english_name_of(entry("1139")).startswith("COATING SOLUTION (")
    assert english_name_is_usable(entry("1139"))


def test_a_reading_cut_off_by_the_column_falls_back_on_the_export():
    """It goes the other way for exactly one entry. UN 2857's name runs past
    the edge of the column and the 2025 reading comes back as "... (UN", where
    the export has "(UN2672)". Preferring the newer edition is not a reason to
    put half a name on a consignment note."""
    from app.services.dg.naming import english_name_is_usable, english_name_of

    from app.services.dg.names_en import english_name as in_the_volume

    assert in_the_volume("2857").endswith("(UN")
    assert english_name_of(entry("2857")).endswith("(UN2672)")
    assert english_name_is_usable(entry("2857"))


def test_an_ordinary_entry_is_not_flagged():
    from app.services.dg.naming import english_name_is_usable

    assert english_name_is_usable(entry(BENZINE))
    # Brackets are ordinary in a shipping name and must not trip the check.
    assert english_name_is_usable(entry("3082"))


def test_the_check_that_says_so_still_refuses_a_name_that_is_not_one():
    """Silence would be the failure: the field is filled, the document looks
    complete, and the name on it is in a language the rulebook does not allow.

    Since v1.89.0 no entry of the table reaches that state — the volume names
    the fourteen the export left empty and the one it cut off — so the check
    is exercised here on what it is for rather than on a UN number that no
    longer needs it. It stays because a later edition can open the gap again."""
    from app.services.dg.naming import english_name_is_usable

    assert not english_name_is_usable({"un": "9999", "name_en": ""})
    assert not english_name_is_usable({"un": "9999", "name_en": "Coating solution ("})
    assert english_name_is_usable({"un": "9999", "name_en": "ACETONE"})


def test_no_entry_of_the_table_needs_that_warning_any_more():
    """The measurement behind the previous test, and the point of reading the
    book: a warning that used to fire on fifteen entries now fires on none."""
    from app.services.dg import database
    from app.services.dg.naming import english_name_is_usable

    unusable = sorted({
        row["un"] for row in database._table_a()
        if (found := database.get_un_entries(row["un"]))
        and not english_name_is_usable(found[0])
    })
    assert unusable == [], unusable


def test_an_entry_with_a_proper_english_name_is_not_reported():
    from app.services.documents.exporter import validate_document
    from app.services.documents.registry import get_document
    from tests.test_documents import BASE_VALUES, LINES

    _errors, warnings = validate_document(
        get_document("cmr"), dict(BASE_VALUES), LINES, german_goods(), "de"
    )
    assert [w for w in warnings if "proper shipping name" in w.lower()
            and "1203" in w] == []


# --- the language of the document, chosen on the export step ---------------
#
# Since v1.76.0 the export asks which language the documents are drawn up in,
# separately from the language of the screen: 5.4.1.4.1 is about the country
# the consignment leaves from, not about who is typing. So a name the app
# derived is re-derived for the document, and only such a name.


def test_de_documenttaal_herschrijft_een_afgeleide_naam():
    product = {"un_number": BENZINE,
               "proper_shipping_name": "BENZINE OF MOTORBRANDSTOF (" + PETROL + ")"}
    assert resolve_for_profile(product, "ADR", "fr")[0] == "ESSENCE"
    assert resolve_for_profile(product, "ADR", "de")[0] == "BENZIN ODER OTTOKRAFTSTOFF"
    assert resolve_for_profile(product, "ADR", "fr")[1] == product["proper_shipping_name"]


def test_zonder_documenttaal_verandert_er_niets():
    product = {"un_number": BENZINE, "proper_shipping_name": "ESSENCE"}
    assert resolve_for_profile(product, "ADR") == ("ESSENCE", "")


def test_eigen_bewoordingen_blijven_staan():
    """A name of the user's own is not one of the four the app can derive, and
    is left exactly as it stands whatever language the document is in."""
    product = {"un_number": BENZINE, "proper_shipping_name": "BENZINE, ONZE EIGEN OMSCHRIJVING"}
    assert resolve_for_profile(product, "ADR", "fr")[0] == "BENZINE, ONZE EIGEN OMSCHRIJVING"


def test_zee_en_lucht_winnen_van_de_taalkeuze():
    product = {"un_number": BENZINE, "proper_shipping_name": "ESSENCE"}
    assert resolve_for_profile(product, "IMDG", "fr")[0] == PETROL


# --- German from the 2025 edition, not the 2023 export ---------------------


def test_de_duitse_naam_komt_uit_de_editie_van_2025():
    """The German names were the last field still coming from a 2023 export.
    The Bundesamt für Strassen edition closed it: 2,346 UN numbers at 0.9996
    agreement, and 2,210 of the names identical to the export that stood
    before — the rest is what two years of ADR did to them."""
    assert proper_shipping_name(entry(BENZINE), "de", ["ADR"]) == "BENZIN ODER OTTOKRAFTSTOFF"
    # The export truncated this one at "Alkohol/Wasser-"; the edition does not.
    assert proper_shipping_name(entry("0219"), "de", ["ADR"]).endswith("MISCHUNG")


def test_een_afbreekstreepje_staat_niet_in_de_naam():
    """The German edition breaks words across the column. A name a driver hands
    over may not carry the typesetter's hyphen: the 2023 export spells this one
    without it too, which is how the reading was checked."""
    assert "-" not in proper_shipping_name(entry("1789"), "de", ["ADR"])
    assert proper_shipping_name(entry("0072"), "de", ["ADR"]).startswith(
        "CYCLOTRIMETHYLENTRINITRAMIN")


def test_een_echt_streepje_blijft_staan():
    """alpha-NAPHTHYLAMIN owns its hyphen: it follows a prefix, not a run of
    capitals."""
    assert proper_shipping_name(entry("2077"), "de", ["ADR"]) == "ALPHA-NAPHTHYLAMIN"
