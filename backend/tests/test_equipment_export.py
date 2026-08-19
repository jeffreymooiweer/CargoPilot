"""The equipment export: the library back out, in the import's own columns.

The point of the export is the round trip — the file it produces is a valid
import, so it is the backup, the hand-over to a colleague who maintains the
list in a spreadsheet, and the seed for the next installation, all in one.
These tests hold it to that: what comes out goes back in and recreates the
library, numbers print as numbers a person can check against reality (never
scientific notation), and nothing is exported unless someone asks — there is
no schedule and no side effect, just a GET.
"""
import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import get_settings
from app.core.database import Base, get_db
from app.core.deps import get_current_user
from app.main import app
from app.models.user import Equipment, User
from app.services.equipment_import import (
    EQUIPMENT_HEADERS,
    _number_cell,
    equipment_to_rows,
    import_equipment_rows,
)


@pytest.fixture
def db(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    db_path = data_dir / "test.db"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    session.add(User(id=1, username="ada", email="ada@example.com", password_hash="x", role="user"))
    session.commit()
    yield session
    session.close()
    get_settings.cache_clear()


@pytest.fixture
def client(db):
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: db.get(User, 1)
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def seed(db):
    db.add(Equipment(
        specifications="DEMO LIGHT VEHICLE", length_cm=400, width_cm=180,
        height_cm=170, wall_thickness_mm=None, weight_kg=1200,
        aliases_json=json.dumps(["demo vehicle", "demo light vehicle"]),
        language_labels_json="{}", source="import", active=True))
    db.add(Equipment(
        specifications="STEEL PLATE 2000x1000x5", length_cm=200, width_cm=100,
        height_cm=0.5, wall_thickness_mm=5, weight_kg=78.5,
        aliases_json="[]", language_labels_json="{}", source="manual", active=False))
    db.commit()


def sheet_rows(content: bytes) -> list[list[str]]:
    ws = load_workbook(io.BytesIO(content)).active
    return [["" if cell is None else str(cell) for cell in row]
            for row in ws.iter_rows(values_only=True)]


# --- the shape of the file ----------------------------------------------------


def test_the_columns_are_the_imports_own(db):
    seed(db)
    rows = equipment_to_rows(db.query(Equipment).order_by(Equipment.specifications).all())
    assert len(rows) == 2
    assert all(len(row) == len(EQUIPMENT_HEADERS) for row in rows)
    vehicle = rows[0]
    assert vehicle[0] == "DEMO LIGHT VEHICLE"
    assert vehicle[1] == "400"
    assert vehicle[6] == "demo vehicle, demo light vehicle"
    assert vehicle[7] == "yes"
    plate = rows[1]
    assert plate[3] == "0.5"
    assert plate[7] == "no"


def test_numbers_print_for_people_not_parsers():
    """1200000 must come out as 1200000 — :g formatting would print 1.2e+06,
    which a parser reads and a person checking the file cannot."""
    assert _number_cell(1_200_000.0) == "1200000"
    assert _number_cell(78.5) == "78.5"
    assert _number_cell(None) == ""


def test_what_comes_out_goes_back_in(db, tmp_path, monkeypatch):
    """The round trip, whole: export the library, import the file into an
    empty installation, and the same records exist there."""
    seed(db)
    exported = equipment_to_rows(db.query(Equipment).order_by(Equipment.specifications).all())

    other_engine = create_engine(
        f"sqlite:///{tmp_path / 'other.db'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=other_engine)
    other = sessionmaker(bind=other_engine)()
    try:
        result = import_equipment_rows(other, [EQUIPMENT_HEADERS] + exported)
        assert result.created == 2
        assert result.errors == []
        plate = other.query(Equipment).filter(
            Equipment.specifications == "STEEL PLATE 2000x1000x5").one()
        assert plate.weight_kg == 78.5
        assert plate.wall_thickness_mm == 5
        assert plate.active is False
        vehicle = other.query(Equipment).filter(
            Equipment.specifications == "DEMO LIGHT VEHICLE").one()
        assert json.loads(vehicle.aliases_json) == ["demo vehicle", "demo light vehicle"]
    finally:
        other.close()


# --- the endpoint ---------------------------------------------------------------


def test_the_export_is_an_xlsx_of_the_whole_library(client, db):
    seed(db)
    response = client.get("/api/equipment/export")
    assert response.status_code == 200
    assert "materieel_export.xlsx" in response.headers["content-disposition"]
    rows = sheet_rows(response.content)
    assert rows[0] == EQUIPMENT_HEADERS
    assert len(rows) == 3
    assert rows[1][0] == "DEMO LIGHT VEHICLE"


def test_an_empty_library_exports_its_headers(client):
    """Headers only — which is exactly the import template, and more honest
    than an error for an installation that simply has not imported yet."""
    response = client.get("/api/equipment/export")
    assert response.status_code == 200
    assert sheet_rows(response.content) == [EQUIPMENT_HEADERS]
