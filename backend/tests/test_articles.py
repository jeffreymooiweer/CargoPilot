"""The articles library: one article per code, found by anything on it,
in and out of a spreadsheet in the same columns, and absent without the
history like everything else the office keeps beyond its accounts.
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import get_settings
from app.core.database import Base, get_db
from app.core.deps import get_current_user
from app.main import create_app
from app.models.user import User
from app.services import articles


@pytest.fixture
def db(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{data_dir / 'test.db'}")
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    monkeypatch.setenv("APP_ENV", "test")
    monkeypatch.setenv("CATALOG_AUTO_SYNC", "false")
    get_settings.cache_clear()
    engine = create_engine(f"sqlite:///{data_dir / 'test.db'}",
                           connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    session.add(User(id=1, username="ada", email="ada@example.com", password_hash="x", role="user"))
    session.commit()
    yield session
    session.close()
    get_settings.cache_clear()


def application(db, monkeypatch, history: bool) -> TestClient:
    monkeypatch.setenv("CARGOPILOT_HISTORY", "true" if history else "false")
    get_settings.cache_clear()
    app = create_app()
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: db.get(User, 1)
    return TestClient(app)


PAINT = {"code": "PAINT-25", "name": "Alkyd paint, 25 L jerrican", "un_number": "UN 1263",
         "proper_shipping_name": "PAINT", "class": "3", "packing_group": "ii",
         "type_of_package": "jerrican", "net_per_package": "25 L"}


def test_no_library_without_the_history(db, monkeypatch):
    with application(db, monkeypatch, history=False) as client:
        assert client.get("/api/articles").status_code == 404
        assert client.post("/api/articles", json=PAINT).status_code == 404


def test_add_find_change_and_remove(db, monkeypatch):
    with application(db, monkeypatch, history=True) as client:
        created = client.post("/api/articles", json=PAINT)
        assert created.status_code == 200, created.text
        article = created.json()
        # The UN number is four digits, the packing group upper case: the
        # library normalises what the office typed in a hurry.
        assert article["un_number"] == "1263" and article["packing_group"] == "II"
        assert article["class"] == "3"
        client.post("/api/articles", json={"code": "BOLT-M12", "name": "Bolts M12, box of 100"})

        assert [a["code"] for a in client.get("/api/articles").json()] == ["BOLT-M12", "PAINT-25"]
        assert [a["code"] for a in client.get("/api/articles?q=1263").json()] == ["PAINT-25"]
        assert [a["code"] for a in client.get("/api/articles?q=bolts").json()] == ["BOLT-M12"]
        assert [a["code"] for a in client.get("/api/articles?q=paint").json()] == ["PAINT-25"]

        changed = client.put(f"/api/articles/{article['id']}", json={**PAINT, "net_per_package": "20 L"})
        assert changed.json()["net_per_package"] == "20 L"
        # A code another article carries is refused on a change.
        assert client.put(f"/api/articles/{article['id']}", json={**PAINT, "code": "bolt-m12"}).status_code == 409

        assert client.delete(f"/api/articles/{article['id']}").json()["ok"] is True
        assert client.delete(f"/api/articles/{article['id']}").status_code == 404


def test_the_same_code_is_one_article(db, monkeypatch):
    with application(db, monkeypatch, history=True) as client:
        first = client.post("/api/articles", json=PAINT).json()
        again = client.post("/api/articles", json={**PAINT, "code": "paint-25", "name": "Paint, new label"}).json()
        assert again["id"] == first["id"] and again["name"] == "Paint, new label"
        assert len(client.get("/api/articles").json()) == 1
        assert client.post("/api/articles", json={"code": "   "}).status_code in (409, 422)


def test_a_code_without_un_number_is_not_dangerous_goods(db):
    article, created = articles.upsert(db, db.get(User, 1), {"code": "BOLT", "un_number": "n/a"})
    assert created and article.un_number == ""
    assert articles.un_digits("un1203") == "1203" and articles.un_digits("12") == "0012"


def _xlsx(rows: list[list[str]]) -> bytes:
    book = Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_a_spreadsheet_goes_in_and_comes_out_in_the_same_columns(db, monkeypatch):
    with application(db, monkeypatch, history=True) as client:
        template = client.get("/api/articles/import-template")
        assert template.status_code == 200
        headers = [c.value for c in load_workbook(io.BytesIO(template.content)).active[1]]
        assert headers == articles.ARTICLE_HEADERS

        # Dutch headers in a different order, a blank code, and a row that
        # brings an existing article up to date.
        client.post("/api/articles", json={"code": "PAINT-25", "name": "old name"})
        rows = [["Artikelcode", "Omschrijving", "UN", "Verpakkingsgroep", "Verpakking", "Inhoud"],
                ["PAINT-25", "Alkyd paint", "UN 1263", "II", "jerrican", "25 L"],
                ["THINNER-5", "Thinner", "1263", "II", "can", "5 L"],
                ["", "no code", "", "", "", ""]]
        response = client.post("/api/articles/import", files={"file": ("articles.xlsx", _xlsx(rows), "application/octet-stream")})
        assert response.status_code == 200, response.text
        assert response.json() == {"created": 1, "updated": 1, "skipped": 1, "errors": []}
        listed = {a["code"]: a for a in client.get("/api/articles").json()}
        assert listed["PAINT-25"]["name"] == "Alkyd paint" and listed["PAINT-25"]["un_number"] == "1263"
        assert listed["THINNER-5"]["net_per_package"] == "5 L"

        exported = client.get("/api/articles/export")
        sheet = load_workbook(io.BytesIO(exported.content)).active
        out = [[c.value for c in row] for row in sheet.iter_rows()]
        assert out[0] == articles.ARTICLE_HEADERS
        assert out[1][:3] == ["PAINT-25", "Alkyd paint", "1263"] and out[1][-1] == "yes"

        # And the export goes straight back in without changing anything.
        again = client.post("/api/articles/import", files={"file": ("articles_export.xlsx", exported.content, "application/octet-stream")})
        assert again.json()["updated"] == 2 and again.json()["created"] == 0
