"""The safety adviser's annual report (ADR 1.8.3.3) over the kept shipments.

What is pinned: the count is over the shipments of one calendar year the
viewer may see, kilograms and litres are kept apart and a quantity without
a unit is unknown rather than guessed, the 1.1.3.6 outcome per shipment is
read from the export as it was kept, the adviser's duties come back as
headings with nothing filled in, and the routes do not exist without the
history.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import get_settings
from app.core.database import Base, get_db
from app.core.deps import get_current_user
from app.main import create_app
from app.models.shipment import Shipment
from app.models.user import Department, User
from app.services import dgsa_report, history
from tests.test_export_bundle import CONSIGNMENT, PRODUCT
from tests.test_history import shipment


@pytest.fixture
def db(tmp_path, monkeypatch):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{data_dir / 'test.db'}")
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    monkeypatch.setenv("APP_ENV", "test")
    monkeypatch.setenv("CATALOG_AUTO_SYNC", "false")
    monkeypatch.setenv("CARGOPILOT_HISTORY", "true")
    get_settings.cache_clear()
    engine = create_engine(f"sqlite:///{data_dir / 'test.db'}",
                           connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    session.add(Department(id=1, name="Sales"))
    session.add_all([
        User(id=1, username="root", email="root@example.com", password_hash="x", role="admin"),
        User(id=2, username="ada", email="ada@example.com", password_hash="x", role="user",
             department_id=1),
    ])
    session.commit()
    yield session
    session.close()
    get_settings.cache_clear()


def client_as(db, user_id: int) -> TestClient:
    app = create_app()
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: db.get(User, user_id)
    return TestClient(app)


def kept(db, user_id: int, when: datetime, products: list[dict] | None, reference: str = "R") -> Shipment:
    dg = [{"line_id": "1", "vehicle": "UNIT-1", "products": products}] if products else []
    payload = history.ShipmentIn(**shipment(
        values={**CONSIGNMENT, "shipment_reference": reference},
        dangerous_goods=dg,
        bundle={"documents": [], "dangerous_goods": dg, "profiles": ["ADR"], "output_language": "nl"}))
    record = history.keep(db, db.get(User, user_id), payload)
    record.created_at = when
    db.commit()
    return record


PETROL = {**PRODUCT, "adr_total_quantity": "800 kg"}            # class 3, PG II
ACETONE = {"un_number": "1090", "proper_shipping_name": "Aceton", "class": "3",
           "packing_group": "II", "transport_category": "2", "adr_total_quantity": "100 L"}
NO_UNIT = {"un_number": "1993", "proper_shipping_name": "Brandbare vloeistof, n.e.g.",
           "class": "3", "packing_group": "III", "transport_category": "3",
           "adr_total_quantity": "5"}
LITHIUM = {"un_number": "3480", "proper_shipping_name": "Lithium-ionbatterijen", "class": "9",
           "packing_group": "", "transport_category": "2", "adr_total_quantity": "12 kg"}


@pytest.fixture
def a_year(db):
    t = lambda m, d: datetime(2026, m, d, 10, 0, tzinfo=timezone.utc)  # noqa: E731
    kept(db, 2, t(1, 5), [PETROL, ACETONE], "S-1")           # Sales: 800 kg + 100 L
    kept(db, 2, t(1, 20), [NO_UNIT], "S-2")                  # Sales: unit unknown
    kept(db, 1, t(6, 1), None, "N-1")                        # nobody's: no DG
    kept(db, 1, t(11, 30), [LITHIUM, PETROL], "N-2")         # nobody's: 12 kg + 800 kg
    kept(db, 1, datetime(2025, 12, 31, 23, 0, tzinfo=timezone.utc), [PETROL], "OLD")


def test_the_years_offered_are_the_years_with_kept_shipments(db, a_year):
    assert dgsa_report.years_kept(db, db.get(User, 1)) == [2026, 2025]
    # Ada sees Sales only, and Sales kept nothing in 2025.
    assert dgsa_report.years_kept(db, db.get(User, 2)) == [2026]


def test_the_administrator_counts_the_whole_year_and_keeps_units_apart(db, a_year):
    report = dgsa_report.build_report(db, db.get(User, 1), 2026, language="en")
    assert report["totals"] == {
        "shipments": 4, "with_dangerous_goods": 3, "without_dangerous_goods": 1,
        "products": 5, "quantity_kg": 1612.0, "quantity_l": 100.0, "quantity_unknown": 1,
    }
    assert report["scope"] == "All departments"
    months = {m["month"]: (m["shipments"], m["with_dangerous_goods"]) for m in report["by_month"]}
    assert months[1] == (2, 2) and months[6] == (1, 0) and months[11] == (1, 1) and months[3] == (0, 0)

    by_class = {c["class"]: c for c in report["by_class"]}
    assert set(by_class) == {"3", "9"}
    assert by_class["3"]["shipments"] == 3
    assert by_class["3"]["products"] == 4
    assert by_class["3"]["quantity_kg"] == 1600.0
    assert by_class["3"]["quantity_l"] == 100.0
    assert by_class["3"]["quantity_unknown"] == 1
    assert by_class["9"] == {"class": "9", "shipments": 1, "products": 1, "quantity_kg": 12.0,
                             "quantity_l": 0.0, "quantity_unknown": 0}

    # The busiest UN number first; the name is the declared one.
    top = report["by_un_number"][0]
    assert (top["un_number"], top["class"], top["packing_group"]) == ("1203", "3", "II")
    assert top["shipments"] == 2 and top["quantity_kg"] == 1600.0
    assert top["name"] == "Benzine"

    departments = {d["department"]: d["shipments"] for d in report["by_department"]}
    assert departments == {"Sales": 2, "Without a department": 2}
    assert report["by_regulation"] == [{"regulation": "ADR", "shipments": 4}]
    assert [m["modality"] for m in report["by_modality"]] == ["road"]
    assert report["by_modality"][0]["label"] == "Road transport"
    # The 1.1.3.6 outcome per shipment, as the export kept it: two exempt
    # (800 kg × 3 = 2400 is above; 12 kg + 800 kg...) — read the statuses
    # rather than recompute them.
    statuses = {row["status"]: row["shipments"] for row in report["adr_points"]}
    assert sum(statuses.values()) == 3
    assert set(statuses) <= {"exempt_possible", "above_threshold", "incomplete", "not_exempt"}
    # The duties are headings, and nothing is filled in for the adviser.
    assert len(report["duties"]) == 13
    assert report["duties"][-1]["key"] == "security_plan"
    assert "1.10.3.2" in report["duties"][-1]["text"]
    assert "1.8.3.3" in report["basis"]


def test_a_user_counts_their_own_departments_shipments_only(db, a_year):
    report = dgsa_report.build_report(db, db.get(User, 2), 2026, language="nl")
    assert report["totals"]["shipments"] == 2
    assert report["totals"]["quantity_kg"] == 800.0
    assert report["scope"] == "Sales"
    # The filter means nothing to them.
    same = dgsa_report.build_report(db, db.get(User, 2), 2026, department="none")
    assert same["totals"]["shipments"] == 2


def test_an_administrator_may_narrow_to_one_department(db, a_year):
    sales = dgsa_report.build_report(db, db.get(User, 1), 2026, department="1", language="de")
    assert sales["totals"]["shipments"] == 2 and sales["scope"] == "Sales"
    nobody = dgsa_report.build_report(db, db.get(User, 1), 2026, department="none", language="fr")
    assert nobody["totals"]["shipments"] == 2 and nobody["scope"] == "Sans service"


def test_an_empty_year_is_an_empty_report_not_an_error(db, a_year):
    report = dgsa_report.build_report(db, db.get(User, 1), 2019)
    assert report["totals"]["shipments"] == 0
    assert report["by_class"] == [] and report["by_un_number"] == []
    assert all(m["shipments"] == 0 for m in report["by_month"])


def test_the_routes_and_the_workbook(db, a_year):
    with client_as(db, 1) as root:
        assert root.get("/api/shipments/report/years").json() == {"years": [2026, 2025]}
        report = root.get("/api/shipments/report?year=2026&language=en").json()
        assert report["totals"]["shipments"] == 4
        assert root.get("/api/shipments/report?year=1999").status_code == 422

        workbook = root.get("/api/shipments/report.xlsx?year=2026&language=en")
        assert workbook.status_code == 200
        assert workbook.headers["content-disposition"].endswith('cargopilot-dgsa-report-2026.xlsx"')
        book = load_workbook(io.BytesIO(workbook.content))
        assert book.sheetnames == ["Summary", "By month", "By mode of transport", "By regulation",
                                   "By department", "By class", "By UN number", "Documents",
                                   "The adviser's duties"]
        summary = {row[0]: row[1] for row in book["Summary"].iter_rows(values_only=True) if row[0]}
        assert summary["Year"] == 2026
        assert summary["Shipments"] == 4
        assert summary["Quantity (kg)"] == 1612.0
        duties = list(book["The adviser's duties"].iter_rows(values_only=True))
        assert duties[0] == ("Duty", "The adviser's finding")
        assert len(duties) == 14
        assert all(row[1] in (None, "") for row in duties[1:])


def test_without_the_history_there_is_no_report(db, monkeypatch):
    monkeypatch.setenv("CARGOPILOT_HISTORY", "false")
    get_settings.cache_clear()
    with client_as(db, 1) as root:
        assert root.get("/api/shipments/report?year=2026").status_code == 404
        assert root.get("/api/shipments/report/years").status_code == 404
